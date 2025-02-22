from typing import Any, Union
from openpyxl import load_workbook, workbook
from openpyxl.workbook.workbook import Workbook
import xlrd
from xlrd.book import Book
from rich.console import Console

console = Console()

XLS = "xls"
XLSX = "xlsx"


class Spreadsheet:
    _wb: Union[Workbook, Book]
    _mode: str

    def __init__(self, path):
        self._read_workbook(path)

    def get_worksheets(self):
        return self._wb.sheetnames if self._mode == XLSX else self._wb.sheet_names()

    def read_sheet(self, sheet: str) -> list[dict[str, Any]]:
        return (
            self._read_xlsx_sheet(sheet)
            if self._mode == XLSX
            else self._read_xls_sheet(sheet)
        )

    def _read_xls_sheet(self, sheet: str):
        worksheet = self._wb.sheet_by_name(sheet)
        return self._map_rows(worksheet.get_rows())

    def _read_xlsx_sheet(self, sheet: str):
        worksheet = self._wb[sheet]
        return self._map_rows(worksheet.rows)

    def _map_rows(self, rows):
        headers = []
        result: list[dict[str, Any]] = list()
        for row in rows:
            if not headers:
                headers = [cell.value for cell in row]
            else:
                mapped_row: dict[str, Any] = dict()
                cell_values = [cell.value for cell in row]
                for i in range(len(cell_values)):
                    header_value = (
                        str(headers[i])
                        if i < len(headers) and headers[i]
                        else f"column_{i}"
                    )
                    mapped_row[header_value] = cell_values[i]
                result.append(mapped_row)
        return result

    def _read_workbook(self, path: str):
        console.log(f"Reading {path}")
        if path.endswith(XLSX):
            self._mode: str = XLSX
            self._wb: Union[Workbook, Book] = load_workbook(path)
        elif path.endswith(XLS):
            console.log("Using legacy mode")
            self._mode: str = XLS
            self._wb: Union[Workbook, Book] = xlrd.open_workbook(path)
