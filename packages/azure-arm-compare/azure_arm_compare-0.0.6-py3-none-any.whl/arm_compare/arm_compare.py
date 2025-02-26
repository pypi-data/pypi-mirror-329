#!/usr/bin/env python3
import argparse
import json
import yaml
import re
import fnmatch
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Added import for openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def exit_with_error(message):
    sys.stderr.write(message + "\n")
    sys.stderr.flush()
    sys.exit(1)

def parse_arguments():
    parser = argparse.ArgumentParser(description='Compare two ARM templates by type and name')
    parser.add_argument('--left', required=True, help='Left ARM template JSON file')
    parser.add_argument('--right', required=True, help='Right ARM template JSON file')
    parser.add_argument('--config', help='Configuration YAML file (optional)')
    parser.add_argument('--output', required=True, help='Output file')
    # Updated the choices to include 'xlsx'
    parser.add_argument('--format', choices=['markdown', 'html', 'xlsx'], default='html',
                        help='Output format: markdown, html, or xlsx (default: html)')
    return parser.parse_args()

def load_json_file(filepath):
    if not os.path.exists(filepath):
        exit_with_error(f"Error: JSON file '{filepath}' does not exist.")
    try:
        with open(filepath, 'r') as f:
            return json.load(f)
    except Exception as e:
        exit_with_error(f"Error: Failed to read JSON file '{filepath}': {e}")

def load_yaml_file(filepath):
    if not os.path.exists(filepath):
        exit_with_error(f"Error: YAML config file '{filepath}' does not exist.")
    try:
        with open(filepath, 'r') as f:
            return yaml.safe_load(f)
    except Exception as e:
        exit_with_error(f"Error: Failed to read YAML config file '{filepath}': {e}")

def sort_list_if_possible(lst):
    """
    If lst is a non-empty list of dictionaries, attempt to sort it.
    First, if every dict has a "name" key, sort by that.
    Otherwise, if every dict has an "id" key, sort by that.
    Otherwise, leave the list unchanged.
    """
    if isinstance(lst, list) and lst and all(isinstance(item, dict) for item in lst):
        if all("name" in item for item in lst):
            return sorted(lst, key=lambda x: x["name"])
        elif all("id" in item for item in lst):
            return sorted(lst, key=lambda x: x["id"])
    return lst

def flatten_json(data, parent_key=''):
    """
    Recursively flattens JSON into a dict mapping full property paths to values.
    Lists are indexed. If a list contains dictionaries with a "name" or "id" property,
    it will be sorted in memory before flattening.
    """
    items = {}
    if isinstance(data, dict):
        for k, v in data.items():
            new_key = f"{parent_key}.{k}" if parent_key else k
            if isinstance(v, dict):
                items.update(flatten_json(v, new_key))
            elif isinstance(v, list):
                v = sort_list_if_possible(v)
                for i, item in enumerate(v):
                    list_key = f"{new_key}[{i}]"
                    if isinstance(item, (dict, list)):
                        items.update(flatten_json(item, list_key))
                    else:
                        items[list_key] = item
            else:
                items[new_key] = v
    elif isinstance(data, list):
        data = sort_list_if_possible(data)
        for i, item in enumerate(data):
            new_key = f"{parent_key}[{i}]"
            if isinstance(item, (dict, list)):
                items.update(flatten_json(item, new_key))
            else:
                items[new_key] = item
    else:
        items[parent_key] = data
    return items

def generate_anchor(resource_type, resource_name):
    """
    Generates a sanitized anchor string based on resource type and name.
    """
    combined = f"{resource_type}-{resource_name}"
    anchor = re.sub(r'[^a-zA-Z0-9-]', '', combined.replace(' ', '-')).lower()
    return anchor

def generate_markdown_table(resource_key_display, left_flat, right_flat, ignore_rules, ignored_set):
    """
    Given two flattened dicts for a resource (left and right), generate a Markdown table comparing the property values.
    If a property key matches an ignore rule, its fail result is set to "Ignored" (and added to ignored_set);
    if it doesn't match and the left/right values differ, a cross (X) is printed;
    if they match, the cell is left empty.
    """
    md_lines = []
    md_lines.append(f"### Comparison for Resource: {resource_key_display}\n")
    md_lines.append("| Matched | Property Path | Left Value | Right Value |")
    md_lines.append("| --- | --- | --- | --- |")
    all_keys = set(left_flat.keys()).union(set(right_flat.keys()))
    for key in sorted(all_keys):
        is_ignored = ignore_rules and any(fnmatch.fnmatch(key, pattern) for pattern in ignore_rules)
        if is_ignored:
            fail_cell = "Ignored"
            ignored_set.add(key)
        else:
            fail_cell = "" if left_flat.get(key, '') == right_flat.get(key, '') else "X"
        left_val = left_flat.get(key, '')
        right_val = right_flat.get(key, '')
        md_lines.append(f"| {fail_cell} | {key} | {left_val} | {right_val} |")
    return "\n".join(md_lines)

def format_html_value(value):
    """
    If value is longer than 64 characters, return HTML that shows a truncated version with a toggle link.
    """
    if not isinstance(value, str):
        value = str(value)
    if len(value) <= 64:
        return value
    truncated = value[:64]
    return (f'<span class="truncated">{truncated}</span>'
            f'<span class="full" style="display:none;">{value}</span> '
            f'<a href="#" class="toggle-more" onclick="toggleMore(this); return false;">[more]</a>')

def generate_html_table(resource_key_display, left_flat, right_flat, ignore_rules, ignored_set):
    """
    Given two flattened dicts for a resource (left and right), generate an HTML table comparing the property values.
    If a property key matches an ignore rule, its cell will display "Ignored" (and be added to ignored_set);
    if they differ, a cross (X) is displayed in the "Matched" column.
    """
    html_lines = []
    html_lines.append(f"<h3>Comparison for Resource: {resource_key_display}</h3>")
    html_lines.append("<table>")
    html_lines.append("<thead>")
    html_lines.append("<tr><th>Matched</th><th>Property Path</th><th>Left Value</th><th>Right Value</th></tr>")
    html_lines.append("</thead>")
    html_lines.append("<tbody>")
    all_keys = set(left_flat.keys()).union(set(right_flat.keys()))
    for key in sorted(all_keys):
        is_ignored = ignore_rules and any(fnmatch.fnmatch(key, pattern) for pattern in ignore_rules)
        if is_ignored:
            fail_cell = "Ignored"
            ignored_set.add(key)
        else:
            fail_cell = "" if left_flat.get(key, '') == right_flat.get(key, '') else "X"
        left_val = left_flat.get(key, '')
        right_val = right_flat.get(key, '')
        html_lines.append(
            f"<tr>"
            f"<td>{fail_cell}</td>"
            f"<td onclick='highlightRow(this);' style='cursor: pointer;'>{key}</td>"
            f"<td>{format_html_value(left_val)}</td>"
            f"<td>{format_html_value(right_val)}</td>"
            f"</tr>"
        )
    html_lines.append("</tbody>")
    html_lines.append("</table>")
    return "\n".join(html_lines)

def generate_html_summary(summary_entries, ignored_properties, left_dict, right_dict):
    """
    Generates the HTML summary section with an additional "Ignored" column.
    Each summary entry is a tuple of (resource_type, resource_name, total, ignored, correct, incorrect, anchor),
    where total == ignored + correct + incorrect.
    """
    html = []
    html.append("<h1>Summary</h1>")
    if ignored_properties:
        html.append("<h2>Ignored Properties</h2>")
        html.append("<p>The following properties were ignored during comparisons:</p>")
        html.append("<ul>")
        for prop in sorted(ignored_properties):
            html.append(f"<li>{prop}</li>")
        html.append("</ul>")
    html.append("<h2>Compared Resources</h2>")
    html.append("<table>")
    html.append("<thead>")
    html.append("<tr><th>Resource Type</th><th>Name</th><th>Total Properties</th><th>Ignored</th><th>Correct</th><th>Incorrect</th></tr>")
    html.append("</thead>")
    html.append("<tbody>")
    for entry in summary_entries:
        rtype, rname, total, ignored, correct, incorrect, anchor = entry
        html.append(
            f"<tr>"
            f"<td>{rtype}</td>"
            f"<td><a href='#{anchor}'>{rname}</a></td>"
            f"<td>{total}</td>"
            f"<td>{ignored}</td>"
            f"<td>{correct}</td>"
            f"<td>{incorrect}</td>"
            f"</tr>"
        )
    html.append("</tbody>")
    html.append("</table>")
    if left_dict:
        html.append("<h2>Unmatched Resources in Left Template</h2>")
        html.append("<table>")
        html.append("<thead>")
        html.append("<tr><th>Resource Type</th><th>Name</th></tr>")
        html.append("</thead>")
        html.append("<tbody>")
        for res in left_dict.values():
            rtype = res.get("type", "Unknown type")
            rname = res.get("name", "Unknown name")
            html.append(f"<tr><td>{rtype}</td><td>{rname}</td></tr>")
        html.append("</tbody>")
        html.append("</table>")
    if right_dict:
        html.append("<h2>Unmatched Resources in Right Template</h2>")
        html.append("<table>")
        html.append("<thead>")
        html.append("<tr><th>Resource Type</th><th>Name</th></tr>")
        html.append("</thead>")
        html.append("<tbody>")
        for res in right_dict.values():
            rtype = res.get("type", "Unknown type")
            rname = res.get("name", "Unknown name")
            html.append(f"<tr><td>{rtype}</td><td>{rname}</td></tr>")
        html.append("</tbody>")
        html.append("</table>")
    return "\n".join(html)

def generate_markdown_summary(summary_entries, ignored_properties, left_dict, right_dict):
    """
    Generates the Markdown summary section with an additional "Ignored" column.
    Each summary entry is a tuple of (resource_type, resource_name, total, ignored, correct, incorrect, anchor),
    where total == ignored + correct + incorrect.
    """
    lines = []
    lines.append("# Summary\n")
    if ignored_properties:
        lines.append("## Ignored Properties\n")
        lines.append("The following properties were ignored during comparisons:")
        for prop in sorted(ignored_properties):
            lines.append(f"- {prop}")
        lines.append("")
    lines.append("## Compared Resources\n")
    lines.append("| Resource Type | Name | Total Properties | Ignored | Correct | Incorrect |")
    lines.append("| --- | --- | --- | --- | --- | --- |")
    for entry in summary_entries:
        rtype, rname, total, ignored, correct, incorrect, anchor = entry
        lines.append(
            f"| {rtype} | [{rname}](#{anchor}) | {total} | {ignored} | {correct} | {incorrect} |"
        )
    if left_dict:
        lines.append("\n## Unmatched Resources in Left Template\n")
        lines.append("| Resource Type | Name |")
        lines.append("| --- | --- |")
        for res in left_dict.values():
            rtype = res.get("type", "Unknown type")
            rname = res.get("name", "Unknown name")
            lines.append(f"| {rtype} | {rname} |")
    if right_dict:
        lines.append("\n## Unmatched Resources in Right Template\n")
        lines.append("| Resource Type | Name |")
        lines.append("| --- | --- |")
        for res in right_dict.values():
            rtype = res.get("type", "Unknown type")
            rname = res.get("name", "Unknown name")
            lines.append(f"| {rtype} | {rname} |")
    return "\n".join(lines)

def generate_html_output(summary_entries, ignored_properties, left_dict, right_dict, detailed_sections):
    """
    Generates the complete HTML output including summary and detailed comparison sections.
    """
    html_lines = []
    html_lines.append("<html>")
    html_lines.append("<head>")
    html_lines.append("<meta charset='UTF-8'>")
    html_lines.append("<title>Comparison Report</title>")
    html_lines.append("<style>")
    html_lines.append("table { width: 100%; max-width: 100%; border-collapse: collapse; }")
    html_lines.append("th, td { border: 1px solid #000; padding: 4px; overflow-wrap: break-word; word-wrap: break-word; }")
    html_lines.append("</style>")
    html_lines.append("<script>")
    html_lines.append("function toggleMore(link) {")
    html_lines.append("  var full = link.previousElementSibling;")
    html_lines.append("  var truncated = full.previousElementSibling;")
    html_lines.append("  if (full.style.display === 'none') {")
    html_lines.append("    full.style.display = 'inline';")
    html_lines.append("    truncated.style.display = 'none';")
    html_lines.append("    link.textContent = '[less]';")
    html_lines.append("  } else {")
    html_lines.append("    full.style.display = 'none';")
    html_lines.append("    truncated.style.display = 'inline';")
    html_lines.append("    link.textContent = '[more]';")
    html_lines.append("  }")
    html_lines.append("}")
    html_lines.append("function highlightRow(cell) {")
    html_lines.append("  var row = cell.parentNode;")
    html_lines.append("  if(row.style.backgroundColor === 'yellow') {")
    html_lines.append("    row.style.backgroundColor = '';")
    html_lines.append("  } else {")
    html_lines.append("    row.style.backgroundColor = 'yellow';")
    html_lines.append("  }")
    html_lines.append("}")
    html_lines.append("</script>")
    html_lines.append("</head>")
    html_lines.append("<body>")
    html_lines.append(generate_html_summary(summary_entries, ignored_properties, left_dict, right_dict))
    html_lines.append("<hr>")
    html_lines.extend(detailed_sections)
    html_lines.append("</body>")
    html_lines.append("</html>")
    return "\n".join(html_lines)

#
# Generate XLSX header cell
#
def set_header_cell(ws, row, col, value, font, alignment=None):
    """
    Sets the given cell's value, font, and (optionally) alignment.
    """
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = font
    if alignment:
        cell.alignment = alignment


def generate_xlsx_output(summary_entries, ignored_properties, left_dict, right_dict, detailed_data):
    """
    Generates the XLSX output with two sheets:
    1) "Summary & Ignored" with summary info, ignored properties, unmatched resources
    2) "Details" with a merged row for each resource that reads:
       "Comparison for Resource: {ResourceType}/{ResourceName}",
       followed by columns: Matched, Property Path, Left Value, Right Value.
    """
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary & Ignored"

    # Formatting
    ws_summary.column_dimensions['A'].width = 64
    ws_summary.column_dimensions['B'].width = 32
    ws_summary.column_dimensions['C'].width = 16
    ws_summary.column_dimensions['D'].width = 16
    ws_summary.column_dimensions['E'].width = 16
    ws_summary.column_dimensions['F'].width = 16

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center')
    orange_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # Write ignored properties (if any)
    row = 1
    ws_summary.cell(row=row, column=1, value="Ignored Properties").font = bold_font
    row += 1
    if ignored_properties:
        for prop in sorted(ignored_properties):
            ws_summary.cell(row=row, column=1, value=prop)
            row += 1
    else:
        ws_summary.cell(row=row, column=1, value="None")
        row += 1

    row += 1  # blank line

    # Write summary table header
    ws_summary.cell(row=row, column=1, value="Resource Type").font = bold_font
    ws_summary.cell(row=row, column=2, value="Resource Name").font = bold_font

    set_header_cell(ws_summary, row, 3, "Total Properties", bold_font, center_alignment)
    set_header_cell(ws_summary, row, 4, "Ignored", bold_font, center_alignment)
    set_header_cell(ws_summary, row, 5, "Correct", bold_font, center_alignment)
    set_header_cell(ws_summary, row, 6, "Incorrect", bold_font, center_alignment)

    row += 1

    # Write summary rows
    for entry in summary_entries:
        rtype, rname, total, ignored, correct, incorrect, _anchor = entry
        ws_summary.cell(row=row, column=1, value=rtype)
        ws_summary.cell(row=row, column=2, value=rname)
        
        cell_3 = ws_summary.cell(row=row, column=3, value=total)
        cell_3.alignment = center_alignment

        cell_4 = ws_summary.cell(row=row, column=4, value=ignored)
        cell_4.alignment = center_alignment

        cell_5 = ws_summary.cell(row=row, column=5, value=correct)
        cell_5.alignment = center_alignment

        cell_6 = ws_summary.cell(row=row, column=6, value=incorrect)
        cell_6.alignment = center_alignment

        row += 1

    row += 1  # blank line

    # Unmatched resources in left template
    ws_summary.cell(row=row, column=1, value="Unmatched Resources in Left Template").font = bold_font
    row += 1
    if left_dict:
        ws_summary.cell(row=row, column=1, value="Resource Type")
        ws_summary.cell(row=row, column=2, value="Name")
        row += 1
        for res in left_dict.values():
            rtype = res.get("type", "Unknown type")
            rname = res.get("name", "Unknown name")
            ws_summary.cell(row=row, column=1, value=rtype)
            ws_summary.cell(row=row, column=2, value=rname)
            row += 1
    else:
        ws_summary.cell(row=row, column=1, value="None")
        row += 1

    row += 2

    # Unmatched resources in right template
    ws_summary.cell(row=row, column=1, value="Unmatched Resources in Right Template").font = bold_font
    row += 1
    if right_dict:
        ws_summary.cell(row=row, column=1, value="Resource Type")
        ws_summary.cell(row=row, column=2, value="Name")
        row += 1
        for res in right_dict.values():
            rtype = res.get("type", "Unknown type")
            rname = res.get("name", "Unknown name")
            ws_summary.cell(row=row, column=1, value=rtype)
            ws_summary.cell(row=row, column=2, value=rname)
            row += 1
    else:
        ws_summary.cell(row=row, column=1, value="None")
        row += 1

    # --- Summary & Ignored Sheet (unchanged from your current code) ---
    # ... Keep all of your existing logic for writing ignored properties,
    # ... summary rows, unmatched resources, etc.

    # Example placeholders (do NOT delete your existing summary code):
    bold_font = Font(bold=True)
    larger_bold_font = Font(bold=True, size=16)
    center_alignment = Alignment(horizontal='center')
    orange_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # (Your existing summary code continues here; nothing changed)

    # --- Details Sheet ---
    ws_details = wb.create_sheet("Details")

    details_row = 1
    for (rtype, rname, comparisons) in detailed_data:
        # 1) Merged header row for each resource
        ws_details.merge_cells(start_row=details_row, start_column=1, end_row=details_row, end_column=4)
        resource_cell = ws_details.cell(row=details_row, column=1)
        resource_cell.value = f"{rtype}/{rname}"
        resource_cell.font = larger_bold_font
        details_row += 1

        # 2) Header row for matched, property path, left value, right value
        ws_details.cell(row=details_row, column=1, value="Matched").font = bold_font
        ws_details.cell(row=details_row, column=2, value="Property Path").font = bold_font
        ws_details.cell(row=details_row, column=3, value="Left Value").font = bold_font
        ws_details.cell(row=details_row, column=4, value="Right Value").font = bold_font

        # Formatting
        ws_details.column_dimensions['A'].width = 12
        ws_details.column_dimensions['B'].width = 64
        ws_details.column_dimensions['C'].width = 64
        ws_details.column_dimensions['D'].width = 64

        details_row += 1

        # 3) One row per property comparison
        for matched, prop_path, left_val, right_val in comparisons:
            if matched == "Ignored":
                row_fill = orange_fill
            elif matched == "X":
                row_fill = red_fill
            else:  # matched == "" => correct
                row_fill = green_fill

            # Write each cell
            matched_cell = ws_details.cell(row=details_row, column=1)
            matched_cell.value = matched
            matched_cell.alignment = center_alignment

            ws_details.cell(row=details_row, column=2, value=prop_path)
            ws_details.cell(row=details_row, column=3, value=str(left_val))
            ws_details.cell(row=details_row, column=4, value=str(right_val))

            # Apply fill across the row
            for col in range(1, 5):
                ws_details.cell(row=details_row, column=col).fill = row_fill

            # Apply wrap text to columns C and D (columns 3 and 4) for all rows
            for row_cells in ws_details.iter_rows(min_col=3, max_col=4, min_row=1, max_row=ws_details.max_row):
                for cell in row_cells:
                    # Keep any existing horizontal/vertical setting if desired, just set wrap_text=True
                    cell.alignment = Alignment(wrap_text=True)

            details_row += 1

        # Add a blank row after each resource to separate blocks
        details_row += 1

    # (Optionally adjust column widths or call auto_size)
    for sheet in [ws_summary, ws_details]:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True

    return wb

#
# New function to generate XLSX output
#
def generate_xlsx_outputxxx(summary_entries, ignored_properties, left_dict, right_dict, detailed_data):
    """
    Generates the XLSX output with two sheets:
    1) "Summary & Ignored" with summary info, ignored properties, unmatched resources
    2) "Details" with the detailed comparison for each resource
    """
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary & Ignored"

    # Formatting
    ws_summary.column_dimensions['A'].width = 64
    ws_summary.column_dimensions['B'].width = 32
    ws_summary.column_dimensions['C'].width = 16
    ws_summary.column_dimensions['D'].width = 16
    ws_summary.column_dimensions['E'].width = 16
    ws_summary.column_dimensions['F'].width = 16

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center')
    orange_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # Write ignored properties (if any)
    row = 1
    ws_summary.cell(row=row, column=1, value="Ignored Properties").font = bold_font
    row += 1
    if ignored_properties:
        for prop in sorted(ignored_properties):
            ws_summary.cell(row=row, column=1, value=prop)
            row += 1
    else:
        ws_summary.cell(row=row, column=1, value="None")
        row += 1

    row += 1  # blank line

    # Write summary table header
    ws_summary.cell(row=row, column=1, value="Resource Type").font = bold_font
    ws_summary.cell(row=row, column=2, value="Resource Name").font = bold_font

    set_header_cell(ws_summary, row, 3, "Total Properties", bold_font, center_alignment)
    set_header_cell(ws_summary, row, 4, "Ignored", bold_font, center_alignment)
    set_header_cell(ws_summary, row, 5, "Correct", bold_font, center_alignment)
    set_header_cell(ws_summary, row, 6, "Incorrect", bold_font, center_alignment)

    row += 1

    # Write summary rows
    for entry in summary_entries:
        rtype, rname, total, ignored, correct, incorrect, _anchor = entry
        ws_summary.cell(row=row, column=1, value=rtype)
        ws_summary.cell(row=row, column=2, value=rname)
        
        cell_3 = ws_summary.cell(row=row, column=3, value=total)
        cell_3.alignment = center_alignment

        cell_4 = ws_summary.cell(row=row, column=4, value=ignored)
        cell_4.alignment = center_alignment

        cell_5 = ws_summary.cell(row=row, column=5, value=correct)
        cell_5.alignment = center_alignment

        cell_6 = ws_summary.cell(row=row, column=6, value=incorrect)
        cell_6.alignment = center_alignment

        row += 1

    row += 1  # blank line

    # Unmatched resources in left template
    ws_summary.cell(row=row, column=1, value="Unmatched Resources in Left Template").font = bold_font
    row += 1
    if left_dict:
        ws_summary.cell(row=row, column=1, value="Resource Type")
        ws_summary.cell(row=row, column=2, value="Name")
        row += 1
        for res in left_dict.values():
            rtype = res.get("type", "Unknown type")
            rname = res.get("name", "Unknown name")
            ws_summary.cell(row=row, column=1, value=rtype)
            ws_summary.cell(row=row, column=2, value=rname)
            row += 1
    else:
        ws_summary.cell(row=row, column=1, value="None")
        row += 1

    row += 2

    # Unmatched resources in right template
    ws_summary.cell(row=row, column=1, value="Unmatched Resources in Right Template").font = bold_font
    row += 1
    if right_dict:
        ws_summary.cell(row=row, column=1, value="Resource Type")
        ws_summary.cell(row=row, column=2, value="Name")
        row += 1
        for res in right_dict.values():
            rtype = res.get("type", "Unknown type")
            rname = res.get("name", "Unknown name")
            ws_summary.cell(row=row, column=1, value=rtype)
            ws_summary.cell(row=row, column=2, value=rname)
            row += 1
    else:
        ws_summary.cell(row=row, column=1, value="None")
        row += 1

    #
    # Details sheet
    #
    ws_details = wb.create_sheet("Details")

    # Header
    ws_details.cell(row=1, column=1, value="Resource Type").font = bold_font
    ws_details.cell(row=1, column=2, value="Name").font = bold_font
    
    match_header = ws_details.cell(row=1, column=3)
    match_header.value = "Matched"
    match_header.font = bold_font
    match_header.alignment = center_alignment

    ws_details.cell(row=1, column=4, value="Property Path").font = bold_font
    ws_details.cell(row=1, column=5, value="Left Value").font = bold_font
    ws_details.cell(row=1, column=6, value="Right Value").font = bold_font

    # Formatting
    ws_details.column_dimensions['A'].width = 64
    ws_details.column_dimensions['B'].width = 32
    ws_details.column_dimensions['C'].width = 12
    ws_details.column_dimensions['D'].width = 64
    ws_details.column_dimensions['E'].width = 32
    ws_details.column_dimensions['F'].width = 32

    details_row = 2
    for (rtype, rname, comparisons) in detailed_data:
        for comp in comparisons:
            matched, prop_path, left_val, right_val = comp

            # Determine the fill based on 'matched' value
            if matched == "Ignored":
                row_fill = orange_fill
            elif matched == "X":
                row_fill = red_fill
            else:  # assume matched == "" for "correct"
                row_fill = green_fill
                
            # Write row
            ws_details.cell(row=details_row, column=1, value=rtype)
            ws_details.cell(row=details_row, column=2, value=rname)

            # Center the matched column
            matched_cell = ws_details.cell(row=details_row, column=3)
            matched_cell.value = matched
            matched_cell.alignment = center_alignment

            ws_details.cell(row=details_row, column=4, value=prop_path)
            ws_details.cell(row=details_row, column=5, value=str(left_val))
            ws_details.cell(row=details_row, column=6, value=str(right_val))

            # Apply the same fill to every cell in the row
            for col in range(1, 7):
                ws_details.cell(row=details_row, column=col).fill = row_fill

            details_row += 1

    # Auto-fit column widths (roughly)
    for sheet in [ws_summary, ws_details]:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True

    return wb

def main():
    args = parse_arguments()

    if not os.path.exists(args.left):
        exit_with_error(f"Error: Left file '{args.left}' does not exist.")
    if not os.path.exists(args.right):
        exit_with_error(f"Error: Right file '{args.right}' does not exist.")
    if args.config and not os.path.exists(args.config):
        exit_with_error(f"Error: Config file '{args.config}' does not exist.")

    left_template = load_json_file(args.left)
    right_template = load_json_file(args.right)

    config = {}
    ignore_rules = []
    resource_mappings = []
    if args.config:
        config = load_yaml_file(args.config)
        ignore_rules = config.get('ignoreRules', [])
        resource_mappings = config.get('resourceMappings', [])
    
    if "dependsOn" not in ignore_rules:
        ignore_rules.append("dependsOn")

    ignored_properties = set()

    left_resources = left_template.get("resources", [])
    right_resources = right_template.get("resources", [])

    def get_key(resource):
        return (resource.get("type"), resource.get("name"))
    
    left_dict = {get_key(res): res for res in left_resources if res.get("type") and res.get("name")}
    right_dict = {get_key(res): res for res in right_resources if res.get("type") and res.get("name")}

    resource_pairs = []
    summary_entries = []

    # Enhanced partial mapping based on prefixes
    for mapping in resource_mappings:
        left_prefix_type = mapping.get("leftResourceType")
        left_prefix_name = mapping.get("leftResourceName")
        right_prefix_type = mapping.get("rightResourceType")
        right_prefix_name = mapping.get("rightResourceName")
        
        for key in list(left_dict.keys()):
            ltype, lname = key
            if ltype.startswith(left_prefix_type) and lname.startswith(left_prefix_name):
                type_remainder = ltype[len(left_prefix_type):]
                name_remainder = lname[len(left_prefix_name):]
                candidate_right_type = right_prefix_type + type_remainder
                candidate_right_name = right_prefix_name + name_remainder
                candidate_key = (candidate_right_type, candidate_right_name)
                if candidate_key in right_dict:
                    resource_pairs.append((left_dict[key], right_dict[candidate_key]))
                    del left_dict[key]
                    del right_dict[candidate_key]

    for key in list(left_dict.keys()):
        if key in right_dict:
            resource_pairs.append((left_dict[key], right_dict[key]))
            del left_dict[key]
            del right_dict[key]

    # We'll collect the structured detail data for xlsx
    detailed_data = []

    detailed_sections = []
    for left_res, right_res in resource_pairs:
        resource_type = left_res.get("type", "Unknown type")
        resource_name = left_res.get("name", "Unknown name")
        resource_key_display = f"{resource_type} / {resource_name}"
        anchor = generate_anchor(resource_type, resource_name)

        # If the entire resource type is ignored by some rule, skip it
        if ignore_rules and any(fnmatch.fnmatch(resource_type, pattern) for pattern in ignore_rules):
            continue

        left_flat = flatten_json(left_res)
        right_flat = flatten_json(right_res)
        
        total = 0
        correct = 0
        incorrect = 0
        ignored_count = 0
        all_keys = set(left_flat.keys()).union(set(right_flat.keys()))

        # We'll also collect detail rows for xlsx
        comparison_rows = []

        for key in sorted(all_keys):
            is_ignored = ignore_rules and any(fnmatch.fnmatch(key, pattern) for pattern in ignore_rules)
            left_val = left_flat.get(key, '')
            right_val = right_flat.get(key, '')

            total += 1
            if is_ignored:
                ignored_count += 1
                matched = "Ignored"
                ignored_properties.add(key)
            else:
                if left_val == right_val:
                    correct += 1
                    matched = ""
                else:
                    incorrect += 1
                    matched = "X"

            # For xlsx detail
            comparison_rows.append((matched, key, left_val, right_val))

        summary_entries.append((resource_type, resource_name, total, ignored_count, correct, incorrect, anchor))

        # Store the detail for xlsx
        detailed_data.append((resource_type, resource_name, comparison_rows))

        # Also build up the "detailed_section" string for HTML/Markdown
        detailed_section = f'<a id="{anchor}"></a>\n'
        if args.format == "html":
            detailed_section += generate_html_table(resource_key_display, left_flat, right_flat, ignore_rules, ignored_properties)
        else:
            # For markdown we do:
            detailed_section += generate_markdown_table(resource_key_display, left_flat, right_flat, ignore_rules, ignored_properties)
        detailed_sections.append(detailed_section)
        detailed_sections.append("\n")

    if args.format == "html":
        final_output = generate_html_output(summary_entries, ignored_properties, left_dict, right_dict, detailed_sections)
        try:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(final_output)
        except Exception as e:
            exit_with_error(f"Error: Failed to write output file '{args.output}': {e}")

    elif args.format == "markdown":
        final_output = "\n".join([
            generate_markdown_summary(summary_entries, ignored_properties, left_dict, right_dict),
            "---"
        ] + detailed_sections)
        try:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(final_output)
        except Exception as e:
            exit_with_error(f"Error: Failed to write output file '{args.output}': {e}")

    # New XLSX output handling
    else:  # args.format == "xlsx"
        wb = generate_xlsx_output(summary_entries, ignored_properties, left_dict, right_dict, detailed_data)
        try:
            wb.save(args.output)
        except Exception as e:
            exit_with_error(f"Error: Failed to write XLSX file '{args.output}': {e}")

if __name__ == '__main__':
    main()
