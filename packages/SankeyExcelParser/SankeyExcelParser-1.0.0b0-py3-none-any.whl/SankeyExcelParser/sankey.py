"""
Auteur : Vincent LE DOZE
Date : 21/04/23
"""

# External libs ---------------------------------------------------------------
import copy
import numpy as np
import pandas as pd
import seaborn as sns

# Local libs ------------------------------------------------------------------
import SankeyExcelParser.io_excel_constants as CONST
import SankeyExcelParser.su_trace as su_trace

# External modules ------------------------------------------------------------
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Local modules ---------------------------------------------------------------
from SankeyExcelParser.sankey_utils.sankey_object import SankeyObject
from SankeyExcelParser.sankey_utils.node import Node
from SankeyExcelParser.sankey_utils.flux import Flux
from SankeyExcelParser.sankey_utils.data import Data
from SankeyExcelParser.sankey_utils.tag_group import TagGroup
from SankeyExcelParser.sankey_utils.functions import _stdStr
from SankeyExcelParser.sankey_utils.functions import _getValueIfPresent
from SankeyExcelParser.sankey_utils.functions import _extractFluxFromMatrix
from SankeyExcelParser.sankey_utils.functions import _createMatrixFromFlux
from SankeyExcelParser.sankey_utils.functions import _reorderTable

# Constants ----------------------------------------------------------
COLOR_WHITE = 'FFFFFF'
COLOR_BLACK = '000000'
COLOR_GREY = 'CFCFCF'
HEADER_ROW_ID = 1  # tables rows start from 1
INDEX_COL_ID = 0  # tables cols start from 0, because fy
SHEET_BY_DEFAULT = {}
SHEET_FORMATING_BY_DEFAULT = {}

# Default excel sheet attributes
SHEET_BY_DEFAULT['name'] = 'Default'  # TODO reprendre nom user
SHEET_BY_DEFAULT['color'] = COLOR_WHITE
SHEET_BY_DEFAULT['table'] = pd.DataFrame()
SHEET_BY_DEFAULT['write_header'] = True
SHEET_BY_DEFAULT['write_index'] = False

# Default excel sheet header's cells formatting
SHEET_FORMATING_BY_DEFAULT['header'] = {}
SHEET_FORMATING_BY_DEFAULT['header']['alignement'] = Alignment(
    horizontal='left',
    vertical='center',
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=False,
    indent=0)
SHEET_FORMATING_BY_DEFAULT['header']['border'] = Border(
    left=Side(border_style="thin", color=COLOR_BLACK),
    right=Side(border_style="thin", color=COLOR_BLACK),
    bottom=Side(border_style="thick", color=COLOR_BLACK))
SHEET_FORMATING_BY_DEFAULT['header']['font'] = Font(
    bold=True,
    color=COLOR_BLACK)
SHEET_FORMATING_BY_DEFAULT['header']['fill'] = PatternFill(
    'solid', fgColor=COLOR_WHITE)
SHEET_FORMATING_BY_DEFAULT['header']['default_height'] = 15
SHEET_FORMATING_BY_DEFAULT['no_header'] = copy.deepcopy(
    SHEET_FORMATING_BY_DEFAULT['header'])

# Default excel sheet header's cells formatting
SHEET_FORMATING_BY_DEFAULT['index'] = copy.deepcopy(
    SHEET_FORMATING_BY_DEFAULT['header'])
SHEET_FORMATING_BY_DEFAULT['index']['default_width'] = 15

# Default excel sheet content's cells formatting
SHEET_FORMATING_BY_DEFAULT['content'] = {}
SHEET_FORMATING_BY_DEFAULT['content']['alignement'] = Alignment(
    horizontal='left',
    vertical='center',
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=False,
    indent=0)
SHEET_FORMATING_BY_DEFAULT['content']['border'] = Border(
    left=Side(border_style="thin", color=COLOR_BLACK),
    right=Side(border_style="thin", color=COLOR_BLACK),
    bottom=Side(border_style="dashed", color=COLOR_BLACK))
SHEET_FORMATING_BY_DEFAULT['content']['font'] = Font(
    bold=False,
    color=COLOR_BLACK)
SHEET_FORMATING_BY_DEFAULT['content']['fill'] = PatternFill(
    'solid', fgColor=COLOR_WHITE)
SHEET_FORMATING_BY_DEFAULT['content']['compute_width'] = True
SHEET_FORMATING_BY_DEFAULT['content']['default_width'] = 15  # Columns default width
SHEET_FORMATING_BY_DEFAULT['content']['default_height'] = 15  # Lines default heights
SHEET_FORMATING_BY_DEFAULT['no_content'] = copy.deepcopy(
    SHEET_FORMATING_BY_DEFAULT['content'])

# If necessary on special cells
# Dict with keys :
# - cols numbers as tuple
# - then rows numbers as tuples
SHEET_FORMATING_BY_DEFAULT['spe_content'] = {}
SHEET_FORMATING_BY_DEFAULT['spe_header'] = {}


# Class ----------------------------------------------------------------------------
class UserExcelConverter(object):
    """
    Allows convertion between the user sheets names, cols names and the
    standard names used when parsing the excel file.

    Attributes
    ----------
    :param sheets_names: TODO
    :type sheets: dict

    :param cols_names: TODO
    :type flux: dict
    """

    def __init__(self, language=CONST.LANG_FR):
        """ Create & Initialize. """
        self._sheets_names = {}
        self._cols_names = {}
        self._language = language

    @property
    def sheets_names(self):
        return self._sheets_names

    def add_new_sheet(self, std_name, user_name):
        if std_name not in self._sheets_names.keys():
            self._sheets_names[std_name] = user_name
            self._cols_names[std_name] = {}

    def get_user_sheet_name(self, std_sheet_name):
        std_sheet_name = self._deal_with_specific_sheet_names(std_sheet_name)
        if std_sheet_name in self._sheets_names.keys():
            return self._sheets_names[std_sheet_name]
        return CONST.DICT_OF_SHEET_NAMES[std_sheet_name][self._language]

    def add_new_col(self, std_sheet_name, std_col_name, user_col_name):
        if std_sheet_name in self._sheets_names.keys():
            if std_col_name not in self._cols_names.keys():
                self._cols_names[std_sheet_name][std_col_name] = user_col_name

    def get_user_col_name(self, std_sheet_name, std_col_name):
        # Standard sheet name
        std_sheet_name = self._deal_with_specific_sheet_names(std_sheet_name)
        # Check if user specified antother name for this col of given sheet
        if std_sheet_name in self._sheets_names.keys():
            if std_col_name in self._cols_names[std_sheet_name].keys():
                return self._cols_names[std_sheet_name][std_col_name]
        # Otherwise, try to give standard col name
        try:
            return CONST.DICT_OF_COLS_NAMES[std_sheet_name][std_col_name][self._language]
        except Exception:
            # Useful for taggroup cols which are not standard cols
            return std_col_name

    def _deal_with_specific_sheet_names(self, sheet_name):
        if sheet_name == CONST.NODE_TYPE_PRODUCT:
            return CONST.PRODUCTS_SHEET
        if sheet_name == CONST.NODE_TYPE_SECTOR:
            return CONST.SECTORS_SHEET
        if sheet_name == CONST.NODE_TYPE_EXCHANGE:
            return CONST.EXCHANGES_SHEET
        return sheet_name


class Sankey(object):
    """
    Contains all Sankey diagram informations.

    Attributes
    ----------
    :param nodes: Dictionnary of all nodes, accessibles with their standardized name.
    :type nodes: dict{node_name=str, node=:class:`sankey.Node`}

    :param flux: Dictionnary of all flux, accessibles with their standardized name.
    :type flux: dict(flux_name=str: flux=:class:`sankey.Flux`)

    :param taggs: Dictionnary of all tags groups, accessible with their type, and then with their
    standardized name.
    :type taggs: dict(tagg_type=str: taggs=dict(tagg_name=str: tagg=:class:`sankey.TagGroup`))

    :param tooltips: All possible tooltips for Sankey objects, with their respective description
    :type tooltips: dict(name=str: description=str)

    :param units: All possible units for Sankey datas, with their respective description
    :type units: dict(name=str: description=str)
    """

    def __init__(self):
        """ Create & Initialize a Sankey object. """
        # Default attributs
        self.nodes = {}
        self.flux = {}
        self.taggs = {
            CONST.TAG_TYPE_DATA: {},
            CONST.TAG_TYPE_FLUX: {},
            CONST.TAG_TYPE_LEVEL: {},
            CONST.TAG_TYPE_NODE: {}
        }
        self.tooltips = {}
        # Attributes linked to reconciliation
        self.units = {}
        self.constraints = {}
        # Allow retreiving user naming convention
        self.xl_user_converter = UserExcelConverter()
        # Other attributes
        self._max_nodes_level = 1
        self.reset_msgs()

    @property
    def max_nodes_level(self):
        return self._max_nodes_level

    @property
    def info_msg(self):
        return self._info_msg

    def add_info_msg(self, msg):
        if type(msg) is str:
            self._info_msg += msg
            # Check last character to be EOL
            if self._info_msg[-1] != '\n':
                self._info_msg += '\n'

    @property
    def warn_msg(self):
        return self._warn_msg

    def add_warn_msg(self, msg):
        if type(msg) is str:
            self._warn_msg += msg
            # Check last character to be EOL
            if self._warn_msg[-1] != '\n':
                self._warn_msg += '\n'

    @property
    def err_msg(self):
        return self._err_msg

    def add_err_msg(self, msg):
        if type(msg) is str:
            self._err_msg += msg
            # Check last character to be EOL
            if self._err_msg[-1] != '\n':
                self._err_msg += '\n'

    def add_msg_on_new_node(self, msg):
        if self._msg_on_new_nodes is None:
            return
        if self._msg_on_new_nodes == 'warn':
            msg += " Not sure if it was intended."
            self.add_warn_msg(msg)
        else:
            self.add_info_msg(msg)

    def add_msg_on_new_flux(self, msg):
        if self._msg_on_new_flux is None:
            return
        if self._msg_on_new_flux == 'warn':
            msg += " Not sure if it was intended."
            self.add_warn_msg(msg)
        else:
            self.add_info_msg(msg)

    def reset_msgs(
        self,
        msg_on_new_nodes=None,
        msg_on_new_flux=None,
    ):
        self._info_msg = ''
        self._warn_msg = ''
        self._err_msg = ''
        self._msg_on_new_nodes = msg_on_new_nodes
        self._msg_on_new_flux = msg_on_new_flux

    def send_msgs(
        self,
        name_of_operation
    ):
        # Send info msgs
        if len(self._info_msg) > 0:
            _ = 'Info(s) on {} :'.format(name_of_operation)
            su_trace.logger.debug(_)
            for _ in self._info_msg.split('\n'):
                if len(_) > 0:
                    su_trace.logger.debug(' - {}'.format(_))
        # Send warning msgs
        if len(self._warn_msg) > 0:
            _ = 'Warning(s) on {} :'.format(name_of_operation)
            su_trace.logger.info(_)
            for _ in self._warn_msg.split('\n'):
                if len(_) > 0:
                    su_trace.logger.info(' - {}'.format(_))
        # Send error msgs
        if len(self._err_msg) > 0:
            _ = 'Error(s) on {} :'.format(name_of_operation)
            su_trace.logger.error(_)
            for _ in self._err_msg.split('\n'):
                if len(_) > 0:
                    su_trace.logger.error(' - {}'.format(_))

    @property
    def data_taggs(self):
        return list(self.taggs[CONST.TAG_TYPE_DATA].values())

    @property
    def flux_taggs(self):
        return list(self.taggs[CONST.TAG_TYPE_FLUX].values())

    @property
    def level_taggs(self):
        return list(self.taggs[CONST.TAG_TYPE_LEVEL].values())

    @property
    def node_taggs(self):
        return list(self.taggs[CONST.TAG_TYPE_NODE].values())

    @property
    def taggs_extra_infos_names(self):
        extra_infos_names = set()
        for taggs_types in self.taggs.values():
            for tagg in taggs_types.values():
                extra_infos_names |= set(tagg.extra_infos_name)
        return sorted(extra_infos_names)

    @property
    def nodes_extra_infos_names(self):
        extra_infos_names = set()
        for node in self.nodes.values():
            extra_infos_names |= set(node.extra_infos_name)
        return sorted(extra_infos_names)

    @property
    def data_extra_infos_names(self):
        extra_infos_names = set()
        for flux in self.flux.values():
            if flux.has_data():
                for data in flux.datas:
                    extra_infos_names |= set(data.extra_infos_name)
        return sorted(extra_infos_names)

    def reset_all_results(self):
        """
        Remove all results datas from Sankey
        Useful when recomputing MFA
        """
        for flux in self.flux.values():
            flux.reset_results()

    def get_tagg_from_name_and_type(self, tagg_name, tagg_type):
        """
        Gives a tagggroup from a name and a type

        Attributes
        ----------
        :param tagg_name: Name of the taggroup
        :type tagg_name: str

        :param tagg_type: Type of the tag group
        :type tagg_type: str

        Returns
        -------
        :return: Taggroup found or None
        :rtype: TagGroup | None
        """
        try:
            tagg_name_ref = _stdStr(tagg_name)
            return self.taggs[tagg_type][tagg_name_ref]
        except Exception:
            return None

    def __repr__(self):
        """
        Gives a string representation of Sankey object.

        Returns
        -------
        :return: String format of self.
        :rtype: str
        """
        s = ''
        # Add nodes
        s += 'Nodes \n'
        s += '-'*40 + '\n'
        for node in self.nodes.values():
            s += '{}\n'.format(node)
        # Add flux
        s += 'Flux \n'
        s += '-'*40 + '\n'
        for flux in self.flux.values():
            s += '{}\n'.format(flux)
        return s

    def update_from_tags_table(
        self,
        table: pd.DataFrame
    ):
        """
        Update self from a tag groups table.

        Exemple of possible tables

        +-----------------+----------+---------------------+------------------+--------+--------+
        | TAG_NAME        | TAG_TYPE | TAG_TAGS            | TAG_COLORS       | INFO 1 | INFO 2 |
        +=================+==========+=====================+==================+========+========+
        | tag g0          | nodeTag  | tag01:tag02:tag03   | hex1:hex2:hex3   |        |        |
        +-----------------+----------+---------------------+------------------+--------+--------+
        | tag g1          | levelTag | tag11:tag12         |                  |        |        |
        +-----------------+----------+---------------------+------------------+--------+--------+
        | tag g1          | levelTag | tag12               | hex1             | special tag |   |
        +-----------------+----------+---------------------+------------------+--------+--------+
        | tag g3 / tag g4 | nodeTag  | tag31:tag32 / tag41 | hex1:hex2 / hex3 |        |        |
        +-----------------+----------+---------------------+------------------+--------+--------+

        Parameters
        ----------
        :param table: Table to parse.
        :type table: panda.DataFrame

        Returns
        -------
        :return: Tuple with boolean at True if everything went ok, False otherwise
                 and an error message if necessary
        :rtype: (bool, str)
        """
        # Init warning message
        self.reset_msgs()
        # Extract columns names that contains extra infos
        taggs_extra_infos = [_ for _ in table.columns if _ not in CONST.TAG_SHEET_COLS]
        # Create new tags from table
        for index in table.index:
            line = table.iloc[index]
            # The / is specific to level tags - it creates antagonists tag groups
            taggs_names = line[CONST.TAG_NAME].split('/')  # List of tag groups' names
            taggs_tags_names = line[CONST.TAG_TAGS].split('/')  # List of tag groups' tags names
            taggs_tags_colors = _getValueIfPresent(line, CONST.TAG_COLOR, None)  # List of tag groups' tags colors
            # We can have no colors
            if (taggs_tags_colors == '') or (taggs_tags_colors is None):
                taggs_tags_colors = [None]*len(taggs_names)
            else:
                taggs_tags_colors = taggs_tags_colors.split('/')
            # If we have antagonist tag grps, do we have the correct number of '/' between tag grps attributes
            if len(taggs_names) != len(taggs_tags_names):
                err = 'At line {} : '.format(index)
                err += 'Not the same amount of separation with \"\\" '
                err += 'for tags for antagonists tags groups \"{}\" '.format(
                    line[CONST.TAG_NAME])
                return False, err
            # If we have antagonist tag grps, check coherence on number of colors attributes
            if len(taggs_names) < len(taggs_tags_colors):
                warn = 'At line {} : '.format(index)
                warn += 'Not the same amount of separation with \"\\" '
                warn += 'for colors for antagonists tags groups \"{}\" '.format(
                        line[CONST.TAG_NAME])
                self.add_warn_msg(warn)
                # Remove surplus colors
                nb_to_pop = len(taggs_tags_colors) - len(taggs_names)
                for _ in range(nb_to_pop):
                    taggs_tags_colors.pop(-1)
            if len(taggs_names) > len(taggs_tags_colors):
                warn = 'At line {} : '.format(index)
                warn += 'Not the same amount of separation with \"\\" '
                warn += 'for colors for antagonists tags groups \"{}\" '.format(
                        line[CONST.TAG_NAME])
                self.add_warn_msg(warn)
                # Complete missing colors
                nb_to_complete = len(taggs_names) - len(taggs_tags_colors)
                for _ in range(nb_to_complete):
                    taggs_tags_colors.append(None)
            # Create tags groups with their respective tags
            prev_taggs = []
            for tagg_name, tagg_tags_names, tagg_tags_colors in zip(taggs_names, taggs_tags_names, taggs_tags_colors):
                # Old tag groups
                if (tagg_name == 'Dimensions'):
                    continue
                # Create new tag groupe
                tagg = self.get_or_create_tagg(tagg_name, line[CONST.TAG_TYPE])
                if tagg is None:
                    err = 'At line {2} : Could not create tag group \"{0}\" : bad type \"{1}\"'.format(
                        line[CONST.TAG_NAME], line[CONST.TAG_TYPE], index)
                    return False, err
                # Add tags and their respective colors to tag groups
                tags_names = tagg_tags_names.split(':')
                if tagg_tags_colors is not None:
                    tags_colors = tagg_tags_colors.split(':')
                    if len(tags_names) > len(tags_colors):
                        tags_colors += [None]*(len(tags_names) - len(tags_colors))
                else:
                    tags_colors = [None]*len(tags_names)
                for tag_name, tag_color in zip(tags_names, tags_colors):
                    tag = tagg.get_or_create_tag(tag_name)
                    tag.color = tag_color
                # Update tag group attributes
                tagg.update(
                    is_palette=_getValueIfPresent(line, CONST.TAG_IS_PALETTE, None),
                    colormap=_getValueIfPresent(line, CONST.TAG_COLORMAP, None))
                # Add tag group extra infos
                for extra_info in taggs_extra_infos:
                    tagg.add_extra_info(extra_info, line[extra_info])
                # If we have antagonists tags, we need to precise it
                for prev_tagg in prev_taggs:
                    tagg.add_antagonist_tagg(prev_tagg)
                prev_taggs.append(tagg)
        return True, self.warn_msg

    def update_from_data_table(
        self,
        input_table: pd.DataFrame,
        warn_on_new_nodes: bool = False,
        warn_on_new_flux: bool = False
    ):
        # Init warning message
        self.reset_msgs(
            msg_on_new_nodes=('warn' if warn_on_new_nodes else None),
            msg_on_new_flux=('warn' if warn_on_new_flux else None))
        # Copy table to avoid modification on reference
        table = input_table.copy()
        # Extra columns in table (more than needed)
        data_extra_infos = \
            set(table.columns) \
            - set(CONST.DATA_SHEET_COLS) \
            - set(self.taggs[CONST.TAG_TYPE_FLUX].keys()) \
            - set(self.taggs[CONST.TAG_TYPE_DATA].keys())
        # Create new flux & data from  table
        for index in table.index:
            # Read line
            line = table.iloc[index]
            # Create Flux
            flux = self.get_or_create_flux(
                line[CONST.DATA_ORIGIN],
                line[CONST.DATA_DESTINATION])
            if flux is None:
                self.add_warn_msg(
                    'At line {} : Could not find or create specified flux.\n'.
                    format(index))
                continue
            # Get dataTag
            ok_data_tags, data_tags = self._get_tags_from_line(CONST.TAG_TYPE_DATA, line)
            # Get fluxTag
            _, flux_tags = self._get_tags_from_line(CONST.TAG_TYPE_FLUX, line)
            # Do not process line if data tags retreiving failed somehow
            if not ok_data_tags:
                self.add_warn_msg(
                    'At line {} : There are problems with the given tags so the line cannot be processed.\n'
                    .format(index))
                continue
            # Corresponding datas in given flux / tags
            datas = flux.get_corresponding_datas_from_tags(data_tags)
            # Read datas attributes
            data_attrs = {
                'value': _getValueIfPresent(line, CONST.DATA_VALUE, None),
                'quantity': _getValueIfPresent(line, CONST.DATA_QUANTITY, None),
                'natural_unit': _getValueIfPresent(line, CONST.DATA_NATURAL_UNIT, None),
                'factor': _getValueIfPresent(line, CONST.DATA_FACTOR, None),
                'sigma_relative': _getValueIfPresent(line, CONST.DATA_UNCERT, CONST.DEFAULT_SIGMA_RELATIVE),
                'source': _getValueIfPresent(line, CONST.DATA_SOURCE, None),
                'hypothesis': _getValueIfPresent(line, CONST.DATA_HYPOTHESIS, None)}
            # Update datas with read attributes & tags infos
            for data in datas:
                # Associated flux tags
                for flux_tag in flux_tags:
                    data.add_tag(flux_tag)
                # Data attributes
                data.update(**data_attrs)
                # Extra infos
                for extra_info in data_extra_infos:
                    data.add_extra_info(
                        extra_info,
                        _getValueIfPresent(line, extra_info, None))
        return True, self.warn_msg

    def update_from_min_max_table(
        self,
        input_table: pd.DataFrame,
        warn_on_new_nodes: bool = False,
        warn_on_new_flux: bool = False
    ):
        # Init warning message
        self.reset_msgs(
            msg_on_new_nodes=('warn' if warn_on_new_nodes else None),
            msg_on_new_flux=('warn' if warn_on_new_flux else None))
        # Copy table to avoid modification on reference
        table = input_table.copy()
        # Create new flux & data from  table
        for index in table.index:
            # Read line
            line = table.iloc[index]
            # Get min max attributes
            min_max_attributes = {
                "min_val": _getValueIfPresent(line, CONST.MIN_MAX_MIN, None),
                "min_quantity": _getValueIfPresent(line, CONST.MIN_MAX_MIN_QUANTITY, None),
                "max_val": _getValueIfPresent(line, CONST.MIN_MAX_MAX, None),
                "max_quantity": _getValueIfPresent(line, CONST.MIN_MAX_MAX_QUANTITY, None)}
            min_max_optionnal_attributes = {
                "natural_unit": _getValueIfPresent(line, CONST.MIN_MAX_NATURAL_UNIT, None),
                "factor": _getValueIfPresent(line, CONST.MIN_MAX_FACTOR, None),
                "hypothesis": _getValueIfPresent(line, CONST.MIN_MAX_HYPOTHESIS, None),
                "source": _getValueIfPresent(line, CONST.MIN_MAX_SOURCE, None)}
            # We create min/max only if we have attributes
            ok_to_parse_min_max = False
            for _ in min_max_attributes.values():
                if (_ is not None):
                    ok_to_parse_min_max = True
                    break
            # We have the necessary attributes to parse
            if ok_to_parse_min_max:
                # First we get or create corresponding flux
                node_orig = line[CONST.MIN_MAX_ORIGIN]
                node_dest = line[CONST.MIN_MAX_DESTINATION]
                corresp_flux = []
                if str(node_orig) == '*':
                    node_dest = _stdStr(node_dest)
                    for flux in self.flux.values():
                        if flux.dest.name == node_dest:
                            corresp_flux.append(flux)
                elif str(node_dest) == '*':
                    node_orig = _stdStr(node_orig)
                    for flux in self.flux.values():
                        if flux.orig.name == node_orig:
                            corresp_flux.append(flux)
                else:
                    _ = self.get_or_create_flux(node_orig, node_dest)
                    if _ is None:
                        self.add_warn_msg(
                            'At line {} : Could not find or create specified flux.\n'.
                            format(index))
                        continue
                    corresp_flux.append(_)
                # Get tags if they exist
                _, flux_tags_present = self._get_tags_from_line(CONST.TAG_TYPE_FLUX, line)
                ok_data_tags, data_tags_present = self._get_tags_from_line(CONST.TAG_TYPE_DATA, line)
                # Do not process line if tags retreiving failed somehow
                if (not ok_data_tags):
                    self.add_warn_msg(
                        'At line {} : There are problems with the given tags so the line cannot be processed.\n'
                        .format(index))
                    continue
                # Update flux or data related to flux
                for flux in corresp_flux:
                    # Corresponding datas
                    datas = flux.get_corresponding_datas_from_tags(data_tags_present)
                    # Update min max for each data
                    for data in datas:
                        for flux_tag in flux_tags_present:
                            data.add_tag(flux_tag)
                        data.min_max.update(**min_max_attributes)
                        data.min_max.update(**min_max_optionnal_attributes)
        return True, self.warn_msg

    def update_from_constraints_table(
        self,
        input_table: pd.DataFrame,
        warn_on_new_nodes: bool = False,
        warn_on_new_flux: bool = False
    ):
        # Init warning message
        self.reset_msgs(
            msg_on_new_nodes=('warn' if warn_on_new_nodes else None),
            msg_on_new_flux=('warn' if warn_on_new_flux else None))
        # Copy table to avoid modification on reference
        table = input_table.copy()
        # Create new flux & data from  table
        for index in table.index:
            # Read line
            line = table.iloc[index]
            # Get or create corresponding flux
            node_orig = line[CONST.CONSTRAINT_ORIGIN]
            node_dest = line[CONST.CONSTRAINT_DESTINATION]
            corresp_flux = []
            if str(node_orig) == '*':
                # TODO
                # A discuter - Dans 'add_other_constraints' de 'SCFMA',
                # avoir '*' signifie tous les noeuds produits existants...
                node_dest = _stdStr(node_dest)
                for flux in self.flux.values():
                    if flux.dest.name == node_dest:
                        corresp_flux.append(flux)
            elif str(node_dest) == '*':
                # TODO
                # A discuter - Dans 'add_other_constraints' de 'SCFMA',
                # avoir '*' signifie tous les noeuds produits existants...
                node_orig = _stdStr(node_orig)
                for flux in self.flux.values():
                    if flux.orig.name == node_orig:
                        corresp_flux.append(flux)
            else:
                _ = self.get_or_create_flux(node_orig, node_dest)
                if _ is None:
                    self.add_warn_msg(
                        'At line {} : Could not find or create specified flux.\n'
                        .format(index))
                    continue
                corresp_flux.append(_)
            # Get corresponding data if it exists
            # ok_flux_tags, flux_tags_present = self._get_tags_from_line(CONST.TAG_TYPE_FLUX, line)
            ok_data_tags, data_tags_present = self._get_tags_from_line(CONST.TAG_TYPE_DATA, line)
            # Do not process line if tags retreiving failed somehow
            if (not ok_data_tags):
                self.add_warn_msg(
                    'At line {} : There are problems with the given tags so the line cannot be processed.\n'
                    .format(index))
                continue
            # Get min max attributes
            try:
                constraint_id = str(_getValueIfPresent(line, CONST.CONSTRAINT_ID, None))
            except Exception:
                err = 'At line {} : Unable to get a valid constraint id.\n'.format(index)
                return False, err
            constraint_attributes = {
                "eq": _getValueIfPresent(line, CONST.CONSTRAINT_EQ, None),
                "ineq_inf": _getValueIfPresent(line, CONST.CONSTRAINT_INEQ_INF, None),
                "ineq_sup": _getValueIfPresent(line, CONST.CONSTRAINT_INEQ_SUP, None),
                "source": _getValueIfPresent(line, CONST.CONSTRAINT_SOURCE, None),
                "hypothesis": _getValueIfPresent(line, CONST.CONSTRAINT_HYPOTHESIS, None),
                "traduction": _getValueIfPresent(line, CONST.CONSTRAINT_TRADUCTION, None)}
            # Update flux or data related to given constraint
            for flux in corresp_flux:
                # Create and add contraint to data
                if len(data_tags_present) > 0:
                    # Find datas that corresponds to given tags
                    datas = flux.get_corresponding_datas_from_tags(
                        data_tags_present)
                    # If multiple datas
                    split_constraint = (len(datas) > 1)
                    # Update constraints for each data
                    for id, data in enumerate(datas):
                        self.add_constraint(
                            '{0}0{1}'.format(constraint_id, id) if split_constraint else constraint_id,
                            data,
                            **constraint_attributes)
                else:  # Create and add contraint to flux
                    self.add_constraint(constraint_id, flux, **constraint_attributes)
        return True, self.warn_msg

    def update_from_result_table(
        self,
        input_table: pd.DataFrame
    ):
        # Init warning message
        self.reset_msgs(
            msg_on_new_nodes='warn',
            msg_on_new_flux='warn')
        # Copy table to avoid modification on reference
        table = input_table.copy()
        # Extra columns in table (more than needed)
        extra_cols = \
            set(table.columns) \
            - set(CONST.RESULTS_SHEET_COLS) \
            - set(self.taggs[CONST.TAG_TYPE_FLUX].keys()) \
            - set(self.taggs[CONST.TAG_TYPE_DATA].keys())
        # Create new flux & data from  table
        for index in table.index:
            # Read line
            line = table.iloc[index]
            # Create Flux
            flux = self.get_or_create_flux(line[CONST.RESULTS_ORIGIN], line[CONST.RESULTS_DESTINATION])
            if flux is None:
                self.add_warn_msg(
                    'At line {} : Could not find or create specified flux.\n'
                    .format(index))
                continue
            # Datatags
            ok_data_tags, data_tags_present = self._get_tags_from_line(CONST.TAG_TYPE_DATA, line)
            # Do not process line if tags retreiving failed somehow
            if not ok_data_tags:
                self.add_warn_msg(
                    'At line {} : There are problems with the given DataTags so the line cannot be processed.\n'
                    .format(index))
                continue
            # Corresponding data - Must find only one or None
            datas = flux.get_corresponding_datas_from_tags(data_tags_present)
            if len(datas) > 1:
                self.add_warn_msg(
                    "At line {} : ".format(index) +
                    "Too much existing & corresponding datas " +
                    "to result with the tags ({}) ".format(
                        ','.join([_.name_unformatted for _ in data_tags_present])) +
                    "We cannot match result with an existing data.\n")
            if len(datas) == 0:
                self.add_warn_msg(
                    "At line {} : ".format(index) +
                    "No data matching with the given result tagged by ({}) ".format(
                        ','.join([_.name_unformatted for _ in data_tags_present])) +
                    "We cannot match result with an existing data.\n")
                continue
            # FluxTags
            _, flux_tags_present = self._get_tags_from_line(CONST.TAG_TYPE_FLUX, line)
            # Read results attributes
            result_attributes = {
                "value": _getValueIfPresent(line, CONST.RESULTS_VALUE, None)}
            # If we have at least one of the columns listed in result_attributes,
            # ... ok to read data
            for _ in result_attributes.values():
                if (_ is not None):
                    result = Data(**result_attributes)
                    result.min_val = _getValueIfPresent(line, CONST.RESULTS_FREE_MIN, None)
                    result.max_val = _getValueIfPresent(line, CONST.RESULTS_FREE_MAX, None)
                    # Link to flux
                    flux.add_result(result)
                    # Apply tags
                    for tag in (data_tags_present + flux_tags_present):
                        result.add_tag(tag)
                    # Set result alterego
                    result.alterego = datas[0]
                    # Save data extra infos
                    for extra_info in extra_cols:
                        result.add_extra_info(extra_info, _getValueIfPresent(line, extra_info, None))
                    # We break the loop
                    break
        return True, self.warn_msg

    def update_from_analysis_table(
        self,
        input_table: pd.DataFrame
    ):
        # Init warning message
        self.reset_msgs(
            msg_on_new_nodes='warn',
            msg_on_new_flux='warn')
        # Copy table to avoid modification on reference
        table = input_table.copy()
        # Extra columns in table (more than needed)
        extra_cols = \
            set(table.columns) \
            - set(CONST.ANALYSIS_SHEET_COLS) \
            - set(self.taggs[CONST.TAG_TYPE_FLUX].keys()) \
            - set(self.taggs[CONST.TAG_TYPE_DATA].keys())
        # Create new flux & data from  table
        for index in table.index:
            # Read line
            line = table.iloc[index]
            # Create Flux
            flux = self.get_or_create_flux(
                line[CONST.RESULTS_ORIGIN],
                line[CONST.RESULTS_DESTINATION])
            if flux is None:
                self.add_warn_msg(
                    'At line {} : Could not find or create specified flux.\n'
                    .format(index))
                continue
            # Data input attributes
            data_attributes = {
                "value": _getValueIfPresent(line, CONST.ANALYSIS_VALUE_IN, None),
                "sigma_relative": _getValueIfPresent(
                    line, CONST.ANALYSIS_VALUE_IN_SIGMA, CONST.DEFAULT_SIGMA_RELATIVE
                ),
                "sigma_percent": _getValueIfPresent(
                    line, CONST.ANALYSIS_VALUE_IN_SIGMA_PRCT, CONST.DEFAULT_SIGMA_PERCENT
                ),
                "min_val": _getValueIfPresent(line, CONST.ANALYSIS_VALUE_MIN_IN, None),
                "max_val": _getValueIfPresent(line, CONST.ANALYSIS_VALUE_MAX_IN, None)}
            # Results attributes
            result_attributes = {
                "value": _getValueIfPresent(line, CONST.RESULTS_VALUE, None)}
            # Analysis attributes
            analysis_attributes = {
                CONST.ANALYSIS_NB_SIGMAS: _getValueIfPresent(line, CONST.ANALYSIS_NB_SIGMAS, None),
                CONST.ANALYSIS_AI: _getValueIfPresent(line, CONST.ANALYSIS_AI, None),
                CONST.ANALYSIS_CLASSIF: _getValueIfPresent(line, CONST.ANALYSIS_CLASSIF, None)}
            # DataTags
            ok_data_tags, data_tags_present = self._get_tags_from_line(CONST.TAG_TYPE_DATA, line)
            # Do not process line if retreiving failed somehow
            if not ok_data_tags:
                self.add_warn_msg(
                    'At line {} : There are problems with the given DataTags so the line cannot be processed.\n'
                    .format(index))
                continue
            # Corresponding data
            datas = flux.get_corresponding_datas_from_tags(data_tags_present)
            # Update data - Must find only one or None
            if len(datas) == 1:
                # Data was read before, we update it
                data = datas[0]
                data.update_unknown_only(**data_attributes)
            else:
                # Too much matching data, error
                self.add_warn_msg(
                    "At line {} : ".format(index) +
                    "Too much existing & corresponding datas " +
                    "with the tags ({}). ".format(
                        ','.join([_.name_unformatted for _ in data_tags_present])) +
                    "We will do our best for matching.\n")
                # Try to match data
                try:
                    # Min dist algo - init
                    value_to_match = float(data_attributes["value"])
                    data = datas[0]
                    min_dist = abs(data.value - value_to_match)
                    # Min dist algo - run
                    for _ in datas[1:]:
                        dist = abs(_.value - value_to_match)
                        if min_dist > dist:
                            min_dist = dist
                            data = _
                    # Update data attributes
                    data.update_unknown_only(**data_attributes)
                # If we have an error, we pass to next line
                except Exception:
                    self.add_warn_msg(
                        " - Couldn't find a matching data.\n")
                    data = None
            # Corresponding result
            results = flux.get_corresponding_results_from_tags(data_tags_present)
            # FluxTags
            _, flux_tags_present = self._get_tags_from_line(CONST.TAG_TYPE_FLUX, line)
            # Update results - Must find only one or None
            result = Data()
            if len(results) == 0:
                # result was not read before, so we update flux
                result = Data(**result_attributes)
                for tag in (data_tags_present + flux_tags_present):
                    result.add_tag(tag)
                flux.add_result(result)
            elif len(results) == 1:
                # result was read before, we update it
                result = results[0]
                result.update_unknown_only(**result_attributes)
                for tag in (flux_tags_present):
                    result.add_tag(tag)
            else:  # Too much matching result, error
                return False, \
                       'At line {} : '.format(index) + \
                       'Result tagged with ({}) appear more than once'.format(
                            ','.join([_.name_unformatted for _ in data_tags_present]))
            # Link result and data together
            result.alterego = data
            # Add analysis attributes
            result.add_extra_infos(analysis_attributes)
            # Add extra info to both
            for extra_info in extra_cols:
                extra_info_value = _getValueIfPresent(line, extra_info, None)
                result.add_extra_info(extra_info, extra_info_value)
        return True, self.warn_msg

    def update_from_uncertainty_table(
        self,
        input_table: pd.DataFrame
    ):
        # Init warning message
        self.reset_msgs(
            msg_on_new_nodes='warn',
            msg_on_new_flux='warn')
        # Copy table to avoid modification on reference
        table = input_table.copy()
        # # Extra columns in table (more than needed)
        # extra_cols = \
        #     set(table.columns) \
        #     - set(CONST.UNCERTAINTY_SHEET_COLS) \
        #     - set(self.taggs[CONST.TAG_TYPE_FLUX].keys()) \
        #     - set(self.taggs[CONST.TAG_TYPE_DATA].keys())
        # Create new flux & data from  table
        for index in table.index:
            # Read line
            line = table.iloc[index]
            # Create Flux
            flux = self.get_or_create_flux(line[CONST.RESULTS_ORIGIN], line[CONST.RESULTS_DESTINATION])
            if flux is None:
                self.add_warn_msg(
                    'At line {} : Could not find or create specified flux.\n'
                    .format(index))
                continue
            # Read monte carlo attributes
            flux.add_monte_carlo(
                _getValueIfPresent(line, CONST.UNCERTAINTY_MC_MU_IN, None),
                _getValueIfPresent(line, CONST.UNCERTAINTY_MC_STD_IN, None),
                _getValueIfPresent(line, CONST.UNCERTAINTY_MC_MU, None),
                _getValueIfPresent(line, CONST.UNCERTAINTY_MC_STD, None),
                _getValueIfPresent(line, CONST.UNCERTAINTY_MC_MIN, None),
                _getValueIfPresent(line, CONST.UNCERTAINTY_MC_MAX, None))
            # Update probas
            for _ in CONST.UNCERTAINTY_PCOLS:
                flux.monte_carlo.add_proba(_, _getValueIfPresent(line, _, None))
            # Update hists
            for _ in CONST.UNCERTAINTY_HCOLS:
                flux.monte_carlo.add_hist(_, _getValueIfPresent(line, _, None))
            # Tags
            ok_flux_tags, flux_tags_present = self._get_tags_from_line(CONST.TAG_TYPE_FLUX, line)
            ok_data_tags, data_tags_present = self._get_tags_from_line(CONST.TAG_TYPE_DATA, line)
            # Do not process line if tags retreiving failed somehow
            if (not ok_data_tags) or (not ok_flux_tags):
                self.add_warn_msg(
                    'At line {} : There are problems with the given tags so the line cannot be processed.\n'
                    .format(index))
                continue
            # Apply tags
            for tag in (flux_tags_present + data_tags_present):
                flux.monte_carlo.add_tag(tag)
        return True, self.warn_msg

    def update_from_conversions_table(
        self,
        input_table: pd.DataFrame,
        node_list: list
    ):
        # Init warning message
        self.reset_msgs(
            msg_on_new_nodes='warn',
            msg_on_new_flux='warn')
        # Copy table to avoid modification on reference
        table = input_table.copy()
        # Extra columns in table (more than needed)
        extra_cols = set(table.columns) - set(CONST.CONVERSIONS_SHEET_COLS)
        # Here, extra infos are tooltips or units descriptions
        # Get tooltip or units descriptions info
        line = table.iloc[0]  # Always the first line
        units = {}
        tooltips = {}
        for extra_info in extra_cols:
            extra_info_value = _getValueIfPresent(line, extra_info, None)
            # If we have "/" in col name, then the col is about unit
            is_extra_info_about_unit = (len(extra_info.split(' / ')) > 1)
            if is_extra_info_about_unit:  # Tooltips
                if 'unité naturelle' in extra_info:
                    extra_info = extra_info.split('/')[1].strip().upper()
                units[extra_info] = extra_info_value
            else:
                tooltips[extra_info] = extra_info_value
        self.units.update(units)
        self.tooltips.update(tooltips)
        # Create new flux & data from  table
        for index in table.index[1:]:  # We pass the first line
            # Read line
            line = table.iloc[index]
            # Find node
            node = self.get_or_create_node(line[CONST.CONVERSIONS_PRODUCT])
            if node is None:
                self.add_warn_msg(
                    'At line {} : Could not find or create specified node.\n'
                    .format(index))
                continue
            node_list.append(node)
            # Unit input attributes
            unit_localisation = node.match_localisation(
                _getValueIfPresent(line, CONST.CONVERSIONS_LOCATION, None))
            node.add_natural_unit(
                _getValueIfPresent(line, CONST.CONVERSIONS_NATURAL_UNIT, None),
                unit_localisation)
            node.add_factor(
                _getValueIfPresent(line, CONST.CONVERSIONS_FACTOR, None),
                unit_localisation)
            # Add tooltips for node
            for tooltip_name, tooltip_description in tooltips.items():
                node.add_tooltip(
                    tooltip_name,
                    tooltip_description,
                    _getValueIfPresent(line, tooltip_name, None))
            # Add other conversions for node
            for unit_name in units.keys():
                factor = _getValueIfPresent(line, unit_name, None)
                if factor is not None:
                    node.add_other_factor(unit_name, factor, unit_localisation)
        return True, self.warn_msg

    def update_from_nodes_table(
        self,
        input_table: pd.DataFrame,
        warn_on_new_nodes: bool = False
    ):
        # Init warning message
        self.reset_msgs(
            msg_on_new_nodes=('warn' if warn_on_new_nodes else None))
        # Copy table to avoid modification on reference table
        table = input_table.copy()
        # Extra columns in table (more than needed)
        extra_cols = \
            set(table.columns) \
            - set(CONST.NODES_SHEET_COLS) \
            - set(self.taggs[CONST.TAG_TYPE_NODE].keys()) \
            - set(self.taggs[CONST.TAG_TYPE_LEVEL].keys())
        # Create new nodes from data available in table
        for index in table.index:
            # Get line
            line = table.iloc[index]
            name = line[CONST.NODES_NODE]
            level = line[CONST.NODES_LEVEL]
            # Do we know this node
            known_node = self.is_node_registered(name)
            # Create node if needed
            node = self.get_or_create_node(name, level)
            if node is None:
                self.add_warn_msg(
                    'At line {} : Could not find or create specified node.\n'
                    .format(index))
                continue
            # If we already knew this node, then next children will be in a new group
            if known_node:
                node.create_new_children_group()
            # Update node
            node.update(
                mat_balance=_getValueIfPresent(line, CONST.NODES_MAT_BALANCE, 1),
                color=_getValueIfPresent(line, CONST.NODES_COLOR, None),
                definition=_getValueIfPresent(line, CONST.NODES_DEFINITIONS, None))
            # Apply node tags
            self._read_tags_from_line(node, CONST.TAG_TYPE_NODE, line)
            # Apply level tags
            self._read_tags_from_line(node, CONST.TAG_TYPE_LEVEL, line)
            # Save extra infos
            for extra_info in extra_cols:
                node.add_extra_info(extra_info, _getValueIfPresent(line, extra_info, None))
            # Does the node have parents ?
            if (level > 1):
                prev_level = level - 1
                for prev_index in reversed(range(index)):
                    prev_line = table.iloc[prev_index]
                    if prev_line[CONST.NODES_LEVEL] == prev_level:
                        parent_name = _stdStr(prev_line[CONST.NODES_NODE])
                        self.nodes[parent_name].add_child(node)
                        break
        return True, self.warn_msg

    def _read_tags_from_line(
        self,
        sankey_object: SankeyObject,
        taggs_type: str,
        line: pd.DataFrame
    ):
        """
        TODO commenter
        """
        for tagg in self.taggs[taggs_type].keys():
            tags_name = _getValueIfPresent(line, tagg, None)
            if tags_name is not None:
                tags_name = str(tags_name).split(':')
                for tag_name in tags_name:
                    tag = self.taggs[taggs_type][tagg].get_tag_from_name(tag_name)
                    if tag is not None:
                        sankey_object.add_tag(tag)
                    else:
                        if (tag_name == '0.') or (tag_name == '0.0'):
                            self.add_warn_msg(
                                'For tag group \"{0}\", tag \"{1}\" does not exist, so not taken in account. '.format(
                                    self.taggs[taggs_type][tagg].name_unformatted,
                                    tag_name
                                )+'Did you mean \"0\" ? \n')
                        else:
                            self.add_warn_msg(
                                'For tag group \"{0}\", tag \"{1}\" does not exist, so not taken in account.\n'.format(
                                    self.taggs[taggs_type][tagg].name_unformatted,
                                    tag_name
                                ))

    def _get_tags_from_line(
        self,
        taggs_type: str,
        line: pd.DataFrame
    ):
        """
        Get all tags related to a given tag type for a given table line.

        Verifications
        -------------
        - Parsing of tags : lone tag as str "tag name" or group of tags as str "tag1:tag2:..."
        - Check if tag group (line header) exists
        - Check if tag exists for given tag group. If not :
            - Warn msg specifiying that given tag will not be taken in account
            - Specific warn msg if tags are 'numbers', only integers are accepted :
                if tag is "1.0" -> trigger a warning msg as it must be "1"

        Parameters
        ----------
        :param taggs_type: Type of tags to find
        :type taggs_type: str

        :param line: Table's line to parse
        :type line: pd.DataFrame

        Returns
        -------
        :return: List of tags that have been found, None if nothing has been found
        :rtype: list | None
        """
        # Tag search is valid
        # 1. If there is no existing tag group for given tag group type
        # 2. If there are tag groups for given tag group type
        #    AND at least one valid tag is given / tag group, considering that no written tag = all tags
        all_tags_search_is_ok = True
        tags_present = []
        for tagg_name in self.taggs[taggs_type].keys():
            # Read value in given line and for tag group's column
            tags_name = _getValueIfPresent(line, tagg_name, None)
            # If we found something for given tag group
            if tags_name is not None:
                tags_search_is_ok = False
                tags_name = str(tags_name).split(':')
                for tag_name in tags_name:
                    tag = self.taggs[taggs_type][tagg_name].get_tag_from_name(tag_name)
                    if tag is not None:
                        tags_present.append(tag)
                        tags_search_is_ok = True
                    else:
                        if (tag_name == '0.') or (tag_name == '0.0'):
                            self.add_warn_msg(
                                'For tag group \"{0}\", tag \"{1}\" is unknown. '.format(
                                    tagg_name, tag_name
                                )+'Did you mean \"0\" ? \n')
                        else:
                            self.add_warn_msg(
                                'For tag group \"{0}\", tag \"{1}\" is unknown.\n'.format(
                                    tagg_name, tag_name
                                ))
                # Search is not ok if no acceptable tag has been found for given tag group
                all_tags_search_is_ok &= tags_search_is_ok
        # Return list of found tags
        return all_tags_search_is_ok, tags_present

    def update_from_matrix_table(
        self,
        table: pd.DataFrame,
        **kwargs
    ):
        # Init options
        warn_on_new_nodes = False  # Warning message on new node creation
        warn_on_new_flux = False  # Warning message on new flux creation
        data_in_matrix = False  # Read data from input matrix
        if 'warn_on_new_nodes' in kwargs.keys():
            warn_on_new_nodes = kwargs['warn_on_new_nodes']
        if 'warn_on_new_flux' in kwargs.keys():
            warn_on_new_flux = kwargs['warn_on_new_flux']
        if 'data_in_matrix' in kwargs.keys():
            data_in_matrix = kwargs['data_in_matrix']
        self.reset_msgs(
            msg_on_new_nodes=('warn' if warn_on_new_nodes else None),
            msg_on_new_flux=('warn' if warn_on_new_flux else None))
        # Check if we will set tags for nodes in rows and in cols
        tagg_name, tagg_type, tag_name_col, tag_name_row = \
            None, None, None, None
        if 'tag_group' in kwargs.keys():
            try:
                tagg_name = kwargs['tags_group']
                tagg_type = kwargs['tags_type']
                tag_name_col = kwargs['tag_col']
                tag_name_row = kwargs['tag_row']
            except Exception:
                err = 'Unable to extract tag group info.'
                return False, err
        # Get or create tag groups if needed, then get corresponding tags
        # for columns and rows
        tagg, tag_col, tag_row = None, None, None
        if tagg_name is not None:
            tagg = self.get_or_create_tagg(tagg_name, tagg_type)
            tag_col = tagg.get_or_create_tag(tag_name_col)
            tag_row = tagg.get_or_create_tag(tag_name_row)
        # Extract list of origins, and destinations
        origins, destinations, values = [], [], []
        _extractFluxFromMatrix(table, origins, destinations, values)
        # Iterate on flux list to create nodes
        for (origin, destination, value) in zip(origins, destinations, values):
            # Will create flux and nodes if they don't exist
            flux = self.get_or_create_flux(origin, destination)
            if flux is None:  # Continue if flux creation has failed
                continue
            # Apply value if needed
            if data_in_matrix and (value is not None):
                for data in flux.datas:
                    data.value = value
            # If we have tag to set for cols and row, we add them to the nodes
            if tag_col is not None:
                self.get_or_create_node(origin).add_tag(tag_col)
            if tag_row is not None:
                self.get_or_create_node(destination).add_tag(tag_row)
        return True, ''

    def autocompute_missing_flux(self):
        """
        Auto compute missing flux.
        """
        # Init output indicators
        all_ok = True
        # Init warning message
        self.reset_msgs(msg_on_new_flux='info')
        # loop on nodes
        for node in self.nodes.values():
            if node.level == 1:
                self._complete_parenthood_flux(node)
                self._complete_single_child_flux(node)
        # Send msgs
        self.send_msgs('Flux completions (based on nodes parenthood relations)')
        # End
        return all_ok

    def _complete_parenthood_flux(self, parent_node):
        """
        Check if flux are consistants throught node parenthood

        Here : Flux from/to child must also exist for parent
        """
        for children_grp in parent_node.children_grps:
            for child_node in children_grp:
                # Firstly, we need to recurse on all children
                # because here we create missing flux for parents
                if child_node.has_at_least_one_child():
                    self._complete_parenthood_flux(child_node)
                # Secondly,
                # If there is a flux from child to another node, then the flux also exists for the parent
                for child_output_flux in child_node.output_flux:
                    child_dest_node = child_output_flux.dest
                    parent_dest_nodes = [parent_output_flux.dest for parent_output_flux in parent_node.output_flux]
                    if child_dest_node not in parent_dest_nodes:
                        self.get_or_create_flux(parent_node.name, child_dest_node.name)
                # Thirdly,
                # If there is a flux to child from another node, then the flux also exists for the parent
                for child_input_flux in child_node.input_flux:
                    child_orig_node = child_input_flux.orig
                    parent_orig_nodes = [parent_input_flux.orig for parent_input_flux in parent_node.input_flux]
                    if child_orig_node not in parent_orig_nodes:
                        self.get_or_create_flux(child_orig_node.name, parent_node.name)

    def _complete_single_child_flux(self, parent_node):
        """
        Check if flux are consistants throught node parenthood

        Here : Flux from/to parent with single child must also exist for the child
        """
        for children_grp in parent_node.children_grps:
            # Firstly, we create missing flux for all children
            # groups with only one child
            if len(children_grp) == 1:
                # Get the unique child
                child_node = children_grp[0]
                # If there is a flux from parent to another node, then the flux also exists for the unique child
                for parent_output_flux in parent_node.output_flux:
                    parent_dest_node = parent_output_flux.dest
                    child_dest_nodes = [child_output_flux.dest for child_output_flux in child_node.output_flux]
                    if parent_dest_node not in child_dest_nodes:
                        self.get_or_create_flux(child_node.name, parent_dest_node.name)
                # If there is a flux to parent from another node, then the flux also exists for the unique child
                for parent_input_flux in parent_node.input_flux:
                    parent_orig_node = parent_input_flux.orig
                    child_orig_nodes = [child_input_flux.orig for child_input_flux in child_node.input_flux]
                    if parent_orig_node not in child_orig_nodes:
                        self.get_or_create_flux(parent_orig_node.name, child_node.name)
            # Secondly, we need to recurse on children
            # because here flux are defined by parents
            for child_node in children_grp:
                if child_node.has_at_least_one_child():
                    self._complete_single_child_flux(child_node)

    def _detect_parenthood_missing_flux(self, parent_node):
        """
        Check if flux are consistants throught node parenthood

        Here : Raise an error if there is a flux from/to given parent
        that does not exist for at least one of its children
        """
        # Init output indicator
        all_ok = True
        # Run
        for children_grp in parent_node.children_grps:
            # If children grp is empty, pass
            if len(children_grp) == 0:
                continue
            # First,
            # Create sets of all incoming & outgoing flux for all children
            # in given children group
            children_input_flux_origs = set()
            children_output_flux_dests = set()
            for child in children_grp:
                for child_input_flux in child.input_flux:
                    children_input_flux_origs.add(child_input_flux.orig)
                for child_output_flux in child.output_flux:
                    children_output_flux_dests.add(child_output_flux.dest)
            # Then,
            # Check all incoming flux for parent
            for parent_input_flux in parent_node.input_flux:
                # Error if any of the incoming flux, does not exist for at least one child
                # in current children group
                if parent_input_flux.orig not in children_input_flux_origs:
                    err_msg = 'Flux inconsistency. '
                    err_msg += 'For node "{}", flux from "{}" does not exists '.format(
                        parent_node.name, parent_input_flux.orig.name)
                    err_msg += 'for at least one of these children nodes ({})'.format(
                        ','.join(['"{}"'.format(_.name) for _ in children_grp]))
                    self.add_err_msg(err_msg)
                    all_ok = False
            # Then,
            # Check all outgoing flux for parent
            for parent_output_flux in parent_node.output_flux:
                # Error if any of the outgoing flux, does not exist for at least one child
                # in current children group
                if parent_output_flux.dest not in children_output_flux_dests:
                    err_msg = 'Flux inconsistency. '
                    err_msg += 'For node "{}", flux to "{}" does not exists '.format(
                        parent_node.name, parent_output_flux.dest.name)
                    err_msg += 'for at least one of these children nodes ({})'.format(
                        ','.join(['"{}"'.format(_.name) for _ in children_grp]))
                    self.add_err_msg(err_msg)
                    all_ok = False
            # Finally, recursion on children
            # because here flux are defined by parents
            for child_node in children_grp:
                if child_node.has_at_least_one_child():
                    all_ok &= self._detect_parenthood_missing_flux(child_node)
        return all_ok

    def autocompute_mat_balance(self):
        """
        Compute matter balance for all nodes. This computation is executed only
        if it was specified by user in input excel file (ie; mat_balance col present)
        """
        # Init logging message
        self.reset_msgs()
        # Get all nodes name for nodes set as origin or destination
        orig = set()
        dest = set()
        for flux in self.flux.values():
            orig.add(flux.orig.name)
            dest.add(flux.dest.name)
        # Find which nodes are set as destination AND origin
        ok_mat_balance = orig & dest
        # Exclude exchange
        try:
            tag_exchange = \
                self.taggs[CONST.TAG_TYPE_NODE][_stdStr(CONST.NODE_TYPE)] \
                    .get_tag_from_name(CONST.NODE_TYPE_EXCHANGE)
            for node in tag_exchange.references:
                node.mat_balance = 0
                # remove node from mat_balance set
                ok_mat_balance -= {node.name}
        except Exception:
            # TODO : erreur sur tag echange qui disparait
            pass
        # Update values if needed
        for node in self.nodes.values():
            # if mat_balance was not correctly set to 1 by the user we correct it and set it to 0
            if node.name in ok_mat_balance:
                if node.mat_balance is None:
                    node.mat_balance = 1
            else:
                if node.mat_balance is not None:
                    if node.mat_balance != 0:
                        msg = 'Node {} : '.format(node.name)
                        msg += 'Matter balance has been changed because it was inconsistent '
                        msg += '(before = {}, now = 0).'.format(node.mat_balance)
                        self.add_info_msg(msg)
                node.mat_balance = 0
        # Send msgs
        self.send_msgs('matter balance autocomputing')

    def has_at_least_one_mat_balance(self):
        for node in self.nodes.values():
            if node.mat_balance is not None:
                return True
        return False

    def autocompute_nodes_levels(self):
        """
        Recompute the Primary level for all node accordingly to global leveling
        """
        for node in self.nodes.values():
            if not node.has_parents():
                node.level = 1
                max_node_level = node.autocompute_level_and_children_levels()
                self._max_nodes_level = max(self._max_nodes_level, max_node_level)

    def check_overall_sankey_structure(self):
        """
        Check if everything is allright regarding sankey structure

        Returns
        -------
        :return: True if everything is ok, False with error message otherwise.
        :rtype: (bool, str)
        """
        # Keep track of errors
        err_msgs = []
        # Check coherence between product, sector and exchanges
        tagg_type_node = self.get_tagg_from_name_and_type(
            CONST.NODE_TYPE,
            CONST.TAG_TYPE_NODE)
        if tagg_type_node is not None:
            # Get Tags
            tag_product = tagg_type_node.get_tag_from_name(CONST.NODE_TYPE_PRODUCT)
            tag_sector = tagg_type_node.get_tag_from_name(CONST.NODE_TYPE_SECTOR)
            tag_exchange = tagg_type_node.get_tag_from_name(CONST.NODE_TYPE_EXCHANGE)
            # Must have product -> sector / exchange
            if tag_product is not None:
                products = tag_product.references
                for product in products:
                    for output_flux in product.output_flux:
                        if output_flux.dest in products:
                            err_msgs.append(
                                'We cannot have this flux {} : reason Product->Product'.format(output_flux))
            # Must have sector -> product / exchange
            if tag_sector is not None:
                sectors = tag_sector.references
                for sector in sectors:
                    for output_flux in sector.output_flux:
                        if output_flux.dest in sectors:
                            err_msgs.append(
                                'We cannot have this flux {} : reason Sector->Sector'.format(output_flux))
            # Must have exchange -> product / exchange
            if tag_exchange is not None:
                exchanges = tag_exchange.references
                for exchange in exchanges:
                    for output_flux in sector.output_flux:
                        if output_flux.dest in exchanges:
                            err_msgs.append(
                                'We cannot have this flux {} : reason Exchange->Exchange'.format(output_flux))
        return len(err_msgs) == 0, '/n'.join(err_msgs)

    def check_overall_sankey_coherence(self):
        """
        Check if everything in sankey is coherent
        """
        # Init logging message
        self.reset_msgs()
        # Init indicator
        all_ok = True
        # Check if there are missing flux in parenthood relations
        for node in self.nodes.values():
            if node.level == 1:
                all_ok &= self._detect_parenthood_missing_flux(node)
        # Check matter balance coherence
        all_ok &= self._check_parenthood_mat_balance_coherence()
        # Check constraints coherence
        all_ok &= self._check_constraints_coherence()
        # Send msgs
        self.send_msgs("Sankey coherence checks")
        # Return
        return all_ok

    def _check_parenthood_mat_balance_coherence(self):
        """
        Check if mat balance are coherents relatively to node parenthood.
        1. If parent has mat_balance at 1, then children can not have mat_balance at 0
        2. If all children has mat_balance at 1, then parent can not have mat_balance at 0
        """
        # Loop on all nodes
        for node in self.nodes.values():
            # Loop on all children grp for given node
            for children_grp in node.children_grps:
                # Protection
                if len(children_grp) == 0:
                    continue
                # Init indicators
                all_children_at_1 = True
                children_at_0 = []
                # Loop on all children
                for child in children_grp:
                    all_children_at_1 &= (child.mat_balance == 1)
                    if child.mat_balance == 0:
                        children_at_0.append(child)
                # Check coherence
                if (node.mat_balance == 1) and (not all_children_at_1):
                    msg = 'Matter balance incoherence. '
                    msg += 'For node "{}", matter balance has been set to be respected (=1) '.format(
                        node.name)
                    msg += 'but not for these following children nodes ({})'.format(
                        ','.join(['"{}"'.format(_.name) for _ in children_at_0]))
                    self.add_warn_msg(msg)
                if (node.mat_balance == 0) and (all_children_at_1):
                    msg = 'Matter balance incoherence. '
                    msg += 'For node "{}", matter balance has been set to be free (=0) '.format(
                        node.name)
                    msg += 'but matter balance has been set to be respected (=1) '
                    msg += 'for all its children nodes ({}) '.format(
                        ','.join(['"{}"'.format(_.name) for _ in children_at_0]))
                    self.add_warn_msg(msg)
        return True

    def _check_constraints_coherence(self):
        """
        Check if constraints are coherents and respect writing conventions.
        """
        ok = True
        for id, constraints in self.constraints.items():
            # List to check datas / flux redondancy
            all_refs = []
            for constraint in constraints:
                ref = constraint.reference
                # Error if ref already present in constraint
                if ref in all_refs:
                    msg = 'Constraint reference repetitions. '
                    msg += 'For constraint with id={} '.format(id)
                    if isinstance(ref, Data):
                        msg += 'Data "{0} -> {1} : {2}", appear more than once. '.format(
                            ref.flux.orig.name,
                            ref.flux.dest.name,
                            [tag.name_unformatted for tag in ref.tags if tag.group.type == CONST.TAG_TYPE_DATA])
                    if isinstance(ref, Flux):
                        msg += 'Flux "{0} -> {1}", appear more than once. '.format(
                            ref.orig.name,
                            ref.dest.name)
                    msg += 'In solving process, it is not possible to process constraints with data redundancies.'
                    self.add_err_msg(msg)
                    # Must send back failure indicator
                    ok = False
                else:
                    all_refs.append(ref)
        # End
        return ok

    def is_node_registered(
        self,
        node
    ):
        """
        Return True if we already have node in our list.

        Parameters
        ----------
        :param node: Name of the node to find or create. Or node object to find.
        :type node: str | Node

        Returns
        -------
        :return: True if the node is in the list, False otherwise.
        :rtype: bool
        """
        if type(node) is str:
            ref_node = _stdStr(node)
            return (ref_node in self.nodes.keys())
        if type(node) is Node:
            return (node in self.nodes.values())

    def get_or_create_node(
        self,
        name: str,
        level: int = 1
    ):
        """
        Return the node with given name.
        - If the node does not exist, we create it.

        Parameters
        ----------
        :param name: Name of the node to find or create.
        :type name: str

        Optional parameters
        -------------------
        :param level: Set level for created node, defaults to 1
        :type level: int, optional

        Returns
        -------
        :return: The node with given name.
        :rtype: :class:`sankey.Node`
        """
        if type(name) is not str:
            return None
        ref_name = _stdStr(name)
        if ref_name not in self.nodes.keys():
            self.nodes[ref_name] = Node(name, level)
            self.add_msg_on_new_node("Created a new node \"{}\".".format(name))
            # Info on maximum of existing levels for nodes
            self._max_nodes_level = max(level, self._max_nodes_level)
        return self.nodes[ref_name]

    def get_or_create_flux(self, orig: str, dest: str):
        """
        Return the flux with given origin and destination nodes.
        - If the flux does not exist, create it.
        - If the nodes do not exist, we create them.

        Parameters
        ----------
        :param orig: Origin node name.
        :type orig: str

        :param dest: Destination node name
        :type dest: str

        Returns
        -------
        :return: The flux between the given origin and destination
        :rtype: :class:`sankey.Flux`
        """
        if (type(orig) is not str) or (type(dest) is not str):
            return None
        ref_name = '{0} - {1}'.format(_stdStr(orig), _stdStr(dest))
        if ref_name not in self.flux.keys():
            # Create nodes if they dont already exist
            node_orig = self.get_or_create_node(orig)
            node_dest = self.get_or_create_node(dest)
            # Create flux
            flux = Flux(node_orig, node_dest)
            # We must instantiate datas for all existing datatags configs
            flux.instanciate_all_datas(
                data_taggs=self.taggs[CONST.TAG_TYPE_DATA])
            # Create reference point in sankey struct
            self.flux[ref_name] = flux
            # Logging message
            msg = "Created a new flux [\"{0}\" -> \"{1}\"].".format(orig, dest)
            self.add_msg_on_new_flux(msg)
        return self.flux[ref_name]

    def get_or_create_tagg(
        self,
        tagg_name: str,
        tagg_type: str,
        tags=''
    ):
        """
        Get tag group related to given name and type.
        Create a new tag group if necessary.

        Parameters
        ----------
        :param tagg_name: Tagg group name
        :type tagg_name: str

        :param tagg_type: Tagg group type
        :type tagg_type: str

        :param tags: tags to add to tag group if newly created
        :type tags: str

        Returns
        -------
        :return: The asked tag group if everything was ok. Else None
        :rtype: TagGroup | None
        """
        # Check if we have the correct type
        if tagg_type not in self.taggs.keys():
            return None
        # Find tag
        ref_tagg_name = _stdStr(tagg_name)
        if ref_tagg_name not in self.taggs[tagg_type].keys():
            self.taggs[tagg_type][ref_tagg_name] = TagGroup(tagg_name, tagg_type, tags=tags)
        return self.taggs[tagg_type][ref_tagg_name]

    def add_constraint(self, id_constraint, reference, **kwargs):
        if isinstance(reference, Flux) or isinstance(reference, Data):
            # Create piece of constraints
            constraint = reference.add_constraint(id_constraint, **kwargs)
            # Update constraint for given id
            if id_constraint in self.constraints.keys():
                self.constraints[id_constraint].append(constraint)
            else:
                self.constraints[id_constraint] = [constraint]

    def has_at_least_one_flux(self):
        return len(self.flux) > 0

    def has_at_least_one_data(self):
        if self.has_at_least_one_flux():
            for flux in self.flux.values():
                if flux.has_data():
                    return True
        return False

    def has_at_least_one_result(self):
        if self.has_at_least_one_flux():
            for flux in self.flux.values():
                if flux.has_result():
                    return True
        return False

    def has_at_least_one_tagg(self):
        has_elems = False
        for tagg_dict in self.taggs.values():
            has_elems |= (len(tagg_dict) > 0)
        return has_elems

    def has_at_least_one_constraint(self):
        return len(self.constraints) > 0

    def write_in_excel_file(
        self,
        excel_file,
        **kwargs
    ):
        """
        _summary_

        Parameters
        ----------
        :param excel_file: Output excel file name
        :type excel_file: file object

        Hidden parameters
        -----------------
        :param additional_sheets:
            Dict of tables as {sheet name as str: table as DataFrame} to add in Excel file
        :type additional_sheets: Dict{str: pandas.DataFrame}
        """
        # Dedicated function to find empty cells
        def is_empty(cell):
            if cell.value is None:
                return True
            if isinstance(cell.value, str):
                if len(cell.value) == 0:
                    return True
            if isinstance(cell.value, int):
                if cell.value == 0:
                    return True
            return False
        # First create sheets as panda tables
        sheets = {}
        nodes_entries = []
        nodes_entries__levels = []
        self.write_tags_sheet(sheets)
        self.write_nodes_sheets(sheets, nodes_entries, nodes_entries__levels)
        self.write_flux_sheets(nodes_entries, nodes_entries__levels, sheets)
        self.write_data_sheets(nodes_entries, sheets)
        # Then write tables in excel file
        for sheet in sheets.values():
            # Dont process empty tables
            if sheet['table'].empty:
                continue
            # Create sheet with data
            sheet['table'].to_excel(
                excel_file,
                sheet_name=sheet['name'],
                index=sheet['write_index'],
                header=sheet['write_header'],
                startrow=0,
                startcol=0)
            # Add formating to sheet
            excel_sheet = excel_file.sheets[sheet['name']]
            excel_sheet.sheet_properties.tabColor = sheet['color']
            # Rows iterator
            rows = excel_sheet.rows
            cols_max_size = []
            # Apply defaut height for header
            excel_sheet.row_dimensions[HEADER_ROW_ID].height = \
                sheet['header']['default_height']
            # Add special formating to header
            header = next(rows)
            for (_, cell) in enumerate(header):
                # Col index
                col_id = INDEX_COL_ID + _  # Because enumerates starts from 0, and table cols from 0
                # Apply different formating depending if cell contains value or not
                if not is_empty(cell):
                    cell.alignment = sheet['header']['alignement']
                    cell.border = sheet['header']['border']
                    cell.fill = sheet['header']['fill']
                    cell.font = sheet['header']['font']
                    if sheet['content']['compute_width']:
                        cols_max_size.append(len(str(cell.value)) + 8)
                    else:
                        cols_max_size.append(sheet['content']['default_width'])
                else:
                    cell.alignment = sheet['no_header']['alignement']
                    cell.border = sheet['no_header']['border']
                    cell.fill = sheet['no_header']['fill']
                    cell.font = sheet['no_header']['font']
                    cols_max_size.append(sheet['content']['default_width'])
                # Apply special formating if needed
                for spe_cols_ids in sheet['spe_header'].keys():
                    if (col_id - INDEX_COL_ID) in spe_cols_ids:  # In special_col_ids, we start from 0
                        cell.alignment = sheet['spe_header'][spe_cols_ids]['alignement']
                        cell.border = sheet['spe_header'][spe_cols_ids]['border']
                        cell.fill = sheet['spe_header'][spe_cols_ids]['fill']
                        cell.font = sheet['spe_header'][spe_cols_ids]['font']
            # Add special formating to the rest of the table
            for (_, row) in enumerate(rows):
                # Row index in table
                row_id = HEADER_ROW_ID + _ + 1  # enumerate starts from 0, but sheet table rows from 1
                # Apply defaut height
                excel_sheet.row_dimensions[row_id].height = \
                    sheet['content']['default_height']
                # Apply formating to each cells
                for (_, cell) in enumerate(row):
                    # Col index in table
                    col_id = INDEX_COL_ID + _  # Because enumerates starts from 0, and table cols from 0
                    # Apply different formating depending if cell contains value or not
                    if not is_empty(cell):
                        # Update cell width from max content
                        if sheet['content']['compute_width']:
                            cols_max_size[col_id] = max(
                                cols_max_size[col_id],
                                len(str(cell.value)) + 8)
                        # Apply content index format if necessary
                        if sheet['write_index'] and (col_id == INDEX_COL_ID):
                            cell.alignment = sheet['index']['alignement']
                            cell.border = sheet['index']['border']
                            cell.fill = sheet['index']['fill']
                            cell.font = sheet['index']['font']
                            cols_max_size[col_id] = sheet['index']['default_width']
                        else:
                            # Apply default content format
                            cell.alignment = sheet['content']['alignement']
                            cell.border = sheet['content']['border']
                            cell.fill = sheet['content']['fill']
                            cell.font = sheet['content']['font']
                    else:
                        cell.alignment = sheet['no_content']['alignement']
                        cell.border = sheet['no_content']['border']
                        cell.fill = sheet['no_content']['fill']
                        cell.font = sheet['no_content']['font']
                    # Apply special formating if needed
                    for spe_cols_ids in sheet['spe_content'].keys():
                        if (col_id - INDEX_COL_ID) in spe_cols_ids:  # In special_col_ids, we start from 0
                            for spe_row_ids in sheet['spe_content'][spe_cols_ids].keys():
                                # /!\ In spe_row_ids, we start from 0
                                if (row_id - HEADER_ROW_ID - 1) in spe_row_ids:
                                    # Force cell formating
                                    cell.alignment = \
                                        sheet['spe_content'][spe_cols_ids][spe_row_ids]['alignement']
                                    cell.border = \
                                        sheet['spe_content'][spe_cols_ids][spe_row_ids]['border']
                                    cell.fill = \
                                        sheet['spe_content'][spe_cols_ids][spe_row_ids]['fill']
                                    cell.font = \
                                        sheet['spe_content'][spe_cols_ids][spe_row_ids]['font']
                                    # Force cell dimension
                                    if 'default_height' in sheet['spe_content'][spe_cols_ids][spe_row_ids].keys():
                                        excel_sheet.row_dimensions[row_id].height = \
                                            sheet['spe_content'][spe_cols_ids][spe_row_ids]['default_height']
            # Apply columns width
            for col_id, col in enumerate(excel_sheet.columns):
                column_letter = col[0].column_letter
                excel_sheet.column_dimensions[column_letter].width = \
                    cols_max_size[col_id]
        # Additionnal sheets
        # Will only work if 'additional_sheets' exists as arg in this function
        try:
            for sheet_name, sheet in kwargs['additional_sheets'].items():
                sheet.to_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    index=False,
                    header=True,
                    startrow=0,
                    startcol=0)
        except Exception:
            pass

    def write_tags_sheet(self, sheets: dict):
        """
        Rewrite tags and taggroups in an excel sheet.

        Parameters
        ----------
        :param sheets: Contains the excel sheets
        :type sheets: dict (output, modified)
        """
        # ----------------------------------------------------
        # Check if we have tags to save
        if not self.has_at_least_one_tagg():
            return
        # ----------------------------------------------------
        # Sheet color
        SHEET_MAIN_COLOR = '9BBB59'
        # Sheet formating infos
        SHEET_FORMATING = copy.deepcopy(SHEET_FORMATING_BY_DEFAULT)
        SHEET_FORMATING['header']['fill'] = PatternFill(
            'solid', fgColor=SHEET_MAIN_COLOR)
        # ----------------------------------------------------
        # Specify columns for table
        taggs_extra_infos_names = self.taggs_extra_infos_names
        table_columns = CONST.TAG_SHEET_COLS + taggs_extra_infos_names
        table_columns = [self.xl_user_converter.get_user_col_name(CONST.TAG_SHEET, _) for _ in table_columns]
        # ----------------------------------------------------
        # Fill table tag with types in specific order
        table_taggs = []
        for tagg_type in [CONST.TAG_TYPE_LEVEL, CONST.TAG_TYPE_NODE, CONST.TAG_TYPE_DATA, CONST.TAG_TYPE_FLUX]:
            antagonists_checked = []
            for tagg in self.taggs[tagg_type].values():
                # Already taken in account as antagonist tagg ?
                if tagg in antagonists_checked:
                    continue
                # Tag groups infos
                name = tagg.name_unformatted
                tags = tagg.tags_str
                # Specific case with antagonist
                if tagg.has_antagonists():
                    for antagonist_tagg in tagg.antagonists_taggs:
                        # One line per pair of antagonist tags
                        antagonist_name = name + '/' + antagonist_tagg.name_unformatted
                        antagonist_tags = tags + '/' + antagonist_tagg.tags_str
                        antagonists_checked.append(antagonist_tagg)
                        # Create table line with corresponding data
                        line_tagg = [
                            antagonist_name,
                            tagg_type,
                            antagonist_tags,
                            tagg.is_palette,
                            tagg.colormap,
                            tagg.colors]
                        # Add extra info cols if needed
                        for extra_info_name in taggs_extra_infos_names:
                            if extra_info_name in tagg.extra_infos.keys():
                                line_tagg.append(tagg.extra_infos[extra_info_name])
                            else:
                                line_tagg.append(None)
                        # We can add it directly in the table
                        table_taggs.append(line_tagg)
                else:
                    # Create table line with corresponding data
                    line_tagg = [
                        name,
                        tagg_type,
                        tags,
                        tagg.is_palette,
                        tagg.colormap,
                        tagg.colors]
                    # Add extra info cols if needed
                    for extra_info_name in taggs_extra_infos_names:
                        if extra_info_name in tagg.extra_infos.keys():
                            line_tagg.append(tagg.extra_infos[extra_info_name])
                        else:
                            line_tagg.append(None)
                    # We can add it directly in the table
                    table_taggs.append(line_tagg)
        table_taggs = pd.DataFrame(table_taggs, columns=table_columns)
        # Drop column that have no values
        table_taggs.dropna(axis=1, how='all', inplace=True)
        # Cast NaN as None because if you have None in a float column,
        # panda transform it as NaN -> cant compare tests after
        table_taggs.replace({np.nan: None}, inplace=True)
        # Update excel sheet attributes
        sheets[CONST.TAG_SHEET] = copy.deepcopy(SHEET_BY_DEFAULT)
        sheets[CONST.TAG_SHEET]['name'] = self.xl_user_converter.get_user_sheet_name(CONST.TAG_SHEET)
        sheets[CONST.TAG_SHEET]['color'] = SHEET_MAIN_COLOR
        sheets[CONST.TAG_SHEET]['table'] = table_taggs
        sheets[CONST.TAG_SHEET].update(SHEET_FORMATING)

    def write_nodes_sheets(
        self,
        sheets: dict,
        nodes_entries: list,
        nodes_entries__levels: list
    ):
        """
        Rewrite nodes and their respective attributes and infos
        in one or somes excel sheets.

        Parameters
        ----------
        :param sheets: Contains the excel sheets
        :type sheets: dict (output, modified)

        :param nodes_entries: List of nodes sorted as they appear in table
        :type nodes_entries: list (output, modified)

        :param nodes_entries__levels: List of levels related to nodes sorted as they appear in table
        :type nodes_entries__levels: list (output, modified)
        """
        # ----------------------------------------------------
        # Sheet color
        SHEET_MAIN_COLOR = '4F81BD'
        # Sheet formating infos
        SHEET_FORMATING = copy.deepcopy(SHEET_FORMATING_BY_DEFAULT)
        SHEET_FORMATING['header']['fill'] = PatternFill(
            'solid', fgColor=SHEET_MAIN_COLOR)
        # Possible types of sheets
        NODES_IN_NODES_SHEET = 1
        NODES_IN_PRODUCTS_SECTORS_EXCHANGES_SHEETS = 2
        # ----------------------------------------------------
        # Default type of sheets
        sheets_type = NODES_IN_NODES_SHEET
        # Columns for tags
        columns_taggs_names = [tagg.name_unformatted for tagg in self.node_taggs]
        columns_taggs_names += [tagg.name_unformatted for tagg in self.level_taggs]
        # If we have node type tag (product:sector:exchange),
        # we remove it from column tags, because
        # we will create 3 tables (product:sector:exchange)
        # instead of only one (nodes)
        if CONST.NODE_TYPE in columns_taggs_names:
            columns_taggs_names.remove(CONST.NODE_TYPE)
            sheets_type = NODES_IN_PRODUCTS_SECTORS_EXCHANGES_SHEETS
        # Specify columns for node table
        nodes_extra_infos_names = self.nodes_extra_infos_names
        table_columns = \
            [CONST.NODES_LEVEL, CONST.NODES_NODE, CONST.NODES_MAT_BALANCE, CONST.NODES_COLOR] + \
            columns_taggs_names + \
            [CONST.NODES_DEFINITIONS] + \
            nodes_extra_infos_names
        # ----------------------------------------------------
        # Fill table node depending on which sheet type
        # - Case 1 : Everything in only one node sheet
        if sheets_type == NODES_IN_NODES_SHEET:
            # Fetch table content line by line
            lineages_tables = []
            lineages_processed = []
            lineages_entries = []
            lineages_entries__levels = []
            for node in self.nodes.values():
                if not node.has_parents():
                    current_lineage_table = []
                    current_lineage_entries = []
                    current_lineage_entries__levels = []
                    lineages_tables.append(current_lineage_table)
                    lineages_entries.append(current_lineage_entries)
                    lineages_entries__levels.append(current_lineage_entries__levels)
                    node.update_table(
                        1,
                        columns_taggs_names,
                        nodes_extra_infos_names,
                        lineages_processed,
                        lineages_tables,
                        current_lineage_table,
                        lineages_entries,
                        current_lineage_entries,
                        lineages_entries__levels,
                        current_lineage_entries__levels)
            # Stack-up all lineage tables
            table_node = sum(lineages_tables, [])
            nodes_entries += sum(lineages_entries, [])
            nodes_entries__levels += sum(lineages_entries__levels, [])
            # Fill table
            table_node = pd.DataFrame(
                table_node,
                columns=[self.xl_user_converter.get_user_col_name(CONST.NODES_SHEET, _) for _ in table_columns])
            # TODO supprimer colonnes vides ou qui contiennent valeurs par défaut
            # Exemple si toutes les couleurs sont à gris
            # Drop column that have no values
            table_node.dropna(axis=1, how='all', inplace=True)
            # Cast NaN as None because if you have None in a float column,
            # panda transform it as NaN -> cant compare tests after
            table_node.replace({np.nan: None}, inplace=True)
            # Save in sheets dictionary
            sheets[CONST.NODES_SHEET] = copy.deepcopy(SHEET_BY_DEFAULT)
            sheets[CONST.NODES_SHEET]['table'] = table_node
            sheets[CONST.NODES_SHEET]['name'] = self.xl_user_converter.get_user_sheet_name(CONST.NODES_SHEET)
            sheets[CONST.NODES_SHEET]['color'] = SHEET_MAIN_COLOR
            sheets[CONST.NODES_SHEET].update(SHEET_FORMATING)
            # ----------------------------------------------------
            # Special formating - Colors for levels
            # Only if we have more than one level
            if self._max_nodes_level <= 1:
                return
            # Crée une palette avec n couleurs celon le nombre de
            # niveau d'agregation allant du bleu clair au blanc
            level_palette = sns \
                .color_palette(
                    "blend:#ffffff,#{}".format(SHEET_MAIN_COLOR),
                    self._max_nodes_level+1) \
                .as_hex()
            # Get colors based on nodes levels
            colors__nodes_indexs = {}
            for index_node, (node, node_level) in enumerate(zip(nodes_entries, nodes_entries__levels)):
                # Color for node level
                color = level_palette[node_level-1].replace('#', '')
                # Update Dict
                if not (color in colors__nodes_indexs.keys()):
                    colors__nodes_indexs[color] = []
                colors__nodes_indexs[color].append(index_node)
            # Save as special formating for excel sheet
            sheets[CONST.NODES_SHEET]['spe_content'][(0, 1)] = {}
            for (color, nodes_indexs) in colors__nodes_indexs.items():
                rows = tuple(nodes_indexs)
                sheets[CONST.NODES_SHEET]['spe_content'][(0, 1)][rows] = \
                    copy.deepcopy(SHEET_FORMATING['content'])
                sheets[CONST.NODES_SHEET]['spe_content'][(0, 1)][rows]['fill'] = \
                    PatternFill('solid', fgColor=color)
        # Case 2 : Nodes are separated in three sheets (Sector, Product, Exchange)
        if sheets_type == NODES_IN_PRODUCTS_SECTORS_EXCHANGES_SHEETS:
            tagg_type_node = self.get_tagg_from_name_and_type(CONST.NODE_TYPE, CONST.TAG_TYPE_NODE)
            for tag in tagg_type_node.tags.values():
                # Fetch content for table
                lineages_tables = []
                lineages_processed = []
                lineages_entries = []
                lineages_entries__levels = []
                for node in tag.references:
                    if not node.has_parents():
                        # Create a new lineage table
                        current_lineage_table = []
                        current_lineage_entries = []
                        current_lineage_entries__levels = []
                        lineages_tables.append(current_lineage_table)
                        lineages_entries.append(current_lineage_entries)
                        lineages_entries__levels.append(current_lineage_entries__levels)
                        # Update given tables
                        node.update_table(
                            1,
                            columns_taggs_names,
                            nodes_extra_infos_names,
                            lineages_processed,
                            lineages_tables,
                            current_lineage_table,
                            lineages_entries,
                            current_lineage_entries,
                            lineages_entries__levels,
                            current_lineage_entries__levels)
                # Stackup tables
                table_node_type = sum(lineages_tables, [])
                nodes_entries__for_tag = sum(lineages_entries, [])
                nodes_entries__levels__for_tag = sum(lineages_entries__levels, [])
                # Fill table
                table_node_type = pd.DataFrame(
                    table_node_type,
                    columns=[self.xl_user_converter.get_user_col_name(tag.name, _) for _ in table_columns])
                # Update node entries
                nodes_entries.extend(nodes_entries__for_tag)
                nodes_entries__levels.extend(nodes_entries__levels__for_tag)
                # TODO supprimer colonnes vides ou qui contiennent valeurs par défaut
                # Drop column that have no values
                table_node_type.dropna(axis=1, how='all', inplace=True)
                # Cast NaN as None because if you have None in a float column,
                # panda transform it as NaN -> cant compare tests after
                table_node_type.replace({np.nan: None}, inplace=True)
                # Save in sheets dictionary
                sheets[tag.name] = copy.deepcopy(SHEET_BY_DEFAULT)
                sheets[tag.name]['table'] = table_node_type
                sheets[tag.name]['name'] = self.xl_user_converter.get_user_sheet_name(tag.name)
                sheets[tag.name]['color'] = SHEET_MAIN_COLOR
                # Update formating with copy, otherwise we have refs values & interference
                # between type node sheets
                sheets[tag.name].update(copy.deepcopy(SHEET_FORMATING))
                # ----------------------------------------------------
                # Special formating - Colors for levels
                # Only if we have more than one level
                if self._max_nodes_level <= 1:
                    continue
                # Crée une palette avec n couleurs celon le nombre de
                # niveau d'agregation allant du bleu clair au blanc
                level_palette = sns \
                    .color_palette(
                        "blend:#ffffff,#{}".format(SHEET_MAIN_COLOR),
                        self._max_nodes_level+1) \
                    .as_hex()
                # Get colors based on nodes levels
                colors__nodes_indexs = {}
                loop_iterator = enumerate(zip(
                    nodes_entries__for_tag,
                    nodes_entries__levels__for_tag))
                for index_node, (node, node_level) in loop_iterator:
                    # Color for node level
                    color = level_palette[node_level-1].replace('#', '')
                    # Update Dict
                    if not (color in colors__nodes_indexs.keys()):
                        colors__nodes_indexs[color] = []
                    colors__nodes_indexs[color].append(index_node)
                # Save as special formating for excel sheet
                sheets[tag.name]['spe_content'][(0, 1)] = {}
                for (color, nodes_indexs) in colors__nodes_indexs.items():
                    rows = tuple(nodes_indexs)
                    sheets[tag.name]['spe_content'][(0, 1)][rows] = \
                        copy.deepcopy(SHEET_FORMATING['content'])
                    sheets[tag.name]['spe_content'][(0, 1)][rows]['fill'] = \
                        PatternFill('solid', fgColor=color)

    def write_flux_sheets(
        self,
        nodes_entries: list,
        nodes_entries__levels: list,
        sheets: dict
    ):
        """
        Rewrite flux and their respective attributes and infos
        in one or some excel sheets.

        Parameters
        ----------
        :param nodes_entries: List of nodes sorted as they appear in nodes tables
        :type nodes_entries: list

        :param nodes_entries__levels: List of levels related to nodes sorted as they appear in nodes tables
        :type nodes_entries__levels: list

        :param sheets: Contains the excel sheets
        :type sheets: dict (output, modified)
        """
        # ----------------------------------------------------
        # Sheet color
        SHEET_MAIN_COLOR = '4F81BD'
        SHEET_CELL_COLOR = '87A9D2'
        # Sheet formating infos
        SHEET_FORMATING = copy.deepcopy(SHEET_FORMATING_BY_DEFAULT)
        SHEET_FORMATING['header']['alignement'] = Alignment(
            horizontal='left',
            vertical='bottom',
            text_rotation=90,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0)
        SHEET_FORMATING['header']['fill'] = PatternFill(
            'solid', fgColor=SHEET_MAIN_COLOR)
        SHEET_FORMATING['header']['border'] = Border(
            right=Side(border_style="dashed", color=COLOR_BLACK),
            bottom=Side(border_style="thick", color=COLOR_BLACK))
        SHEET_FORMATING['no_header']['border'] = Border(
            right=Side(border_style="thin", color=COLOR_BLACK),
            bottom=Side(border_style="thin", color=COLOR_BLACK))
        SHEET_FORMATING['index']['alignement'] = Alignment(
            horizontal='right',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=1)
        SHEET_FORMATING['index']['fill'] = PatternFill(
            'solid', fgColor=SHEET_MAIN_COLOR)
        SHEET_FORMATING['index']['border'] = Border(
            right=Side(border_style="thick", color=COLOR_BLACK),
            bottom=Side(border_style="dashed", color=COLOR_BLACK))
        SHEET_FORMATING['content']['alignement'] = Alignment(
            horizontal='center',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0)
        SHEET_FORMATING['content']['fill'] = PatternFill(
            'solid', fgColor=SHEET_CELL_COLOR)
        SHEET_FORMATING['content']['border'] = Border(
            left=Side(border_style="thin", color=COLOR_BLACK),
            right=Side(border_style="thin", color=COLOR_BLACK),
            top=Side(border_style="thin", color=COLOR_BLACK),
            bottom=Side(border_style="thin", color=COLOR_BLACK))
        SHEET_FORMATING['no_content']['fill'] = PatternFill(
            'solid', fgColor=COLOR_GREY)
        SHEET_FORMATING['no_content']['border'] = Border(
            left=Side(border_style="none"),
            right=Side(border_style="none"),
            bottom=Side(border_style="none"))
        # Possible types of sheets
        FLUX_IN_IO_SHEET = 1
        FLUX_IN_TER_SHEETS = 2
        # ----------------------------------------------------
        # Default type of sheets
        sheets_type = FLUX_IN_IO_SHEET
        tagg_type_node = self.get_tagg_from_name_and_type(
            CONST.NODE_TYPE,
            CONST.TAG_TYPE_NODE)
        if tagg_type_node is not None:
            has_product_tagged_nodes = len(tagg_type_node.tags[CONST.NODE_TYPE_PRODUCT].references) > 0
            has_sector_tagged_nodes = len(tagg_type_node.tags[CONST.NODE_TYPE_SECTOR].references) > 0
            has_exchange_tagged_nodes = len(tagg_type_node.tags[CONST.NODE_TYPE_EXCHANGE].references) > 0
            ok_for_ter_matrix = \
                (has_product_tagged_nodes or has_exchange_tagged_nodes) and \
                (has_sector_tagged_nodes or has_exchange_tagged_nodes)
            if ok_for_ter_matrix:
                sheets_type = FLUX_IN_TER_SHEETS
        # ----------------------------------------------------
        # Fill table node depending on which sheet type
        # - Case 1 : Everything in only one IO sheet
        if sheets_type == FLUX_IN_IO_SHEET:
            # Create matrix
            matrix = _createMatrixFromFlux(
                nodes_entries,
                nodes_entries)
            # From matrix create table with correct header & index names
            nodes_entries_names = [_.name for _ in nodes_entries]
            table = pd.DataFrame(
                matrix,
                index=nodes_entries_names,
                columns=nodes_entries_names)
            # Save in sheets table and parameters
            sheets[CONST.IO_SHEET] = copy.deepcopy(SHEET_BY_DEFAULT)
            sheets[CONST.IO_SHEET]['table'] = table
            sheets[CONST.IO_SHEET]['name'] = self.xl_user_converter.get_user_sheet_name(CONST.IO_SHEET)
            sheets[CONST.IO_SHEET]['color'] = SHEET_MAIN_COLOR
            sheets[CONST.IO_SHEET]['write_index'] = True
            # Initialize default values for sheet formating
            sheets[CONST.IO_SHEET].update(copy.deepcopy(SHEET_FORMATING))
            # For index and header, the height and width depend on nodes names
            max_size = len(max(nodes_entries_names)) + 10
            sheets[CONST.IO_SHEET]['header']['default_height'] = max_size*10
            sheets[CONST.IO_SHEET]['index']['default_width'] = max_size*2
            # For content, use fixed width and heigh
            sheets[CONST.IO_SHEET]['content']['compute_width'] = False
            sheets[CONST.IO_SHEET]['content']['default_width'] = 5
            sheets[CONST.IO_SHEET]['content']['default_height'] = 15
            # ----------------------------------------------------
            # Special formating - Colors for levels
            # Only if we have more than one level
            if self._max_nodes_level <= 1:
                return
            # Create a color gradient (from white to main_color) to fill nodes
            # cells depending on their respective level
            level_palette = sns \
                .color_palette(
                    "blend:#ffffff,#{}".format(SHEET_MAIN_COLOR),
                    self._max_nodes_level+1) \
                .as_hex()
            # Get colors based on nodes levels
            colors__nodes_indexs = {}
            for index_node, (node, node_level) in enumerate(zip(nodes_entries, nodes_entries__levels)):
                # Color for node level
                color = level_palette[node_level-1].replace('#', '')
                # Update Dict
                if not (color in colors__nodes_indexs.keys()):
                    colors__nodes_indexs[color] = []
                colors__nodes_indexs[color].append(index_node)
            # Save as special formating for excel sheet
            sheets[CONST.IO_SHEET]['spe_content'][(0,)] = {}
            for (color, nodes_indexs) in colors__nodes_indexs.items():
                # For header, shift one col to the right, because we have the index col in first col
                cols = tuple(_+1 for _ in nodes_indexs)
                # For index no shift
                rows = tuple(nodes_indexs)
                # First row (Header)
                sheets[CONST.IO_SHEET]['spe_header'][cols] = \
                    copy.deepcopy(sheets[CONST.IO_SHEET]['header'])  # Keep other header's formating attributes
                sheets[CONST.IO_SHEET]['spe_header'][cols]['fill'] = \
                    PatternFill('solid', fgColor=color)  # Apply color filling based on node's level
                # First col (Index)
                sheets[CONST.IO_SHEET]['spe_content'][(0,)][rows] = \
                    copy.deepcopy(sheets[CONST.IO_SHEET]['index'])  # Keep other index's formating attributes
                sheets[CONST.IO_SHEET]['spe_content'][(0,)][rows]['fill'] = \
                    PatternFill('solid', fgColor=color)  # Apply color filling based on node's level
        # - Case 2 : Everything in only one TER sheet
        if sheets_type == FLUX_IN_TER_SHEETS:
            # Number of rows between the two matrixs
            NB_ROWS_BETWEEN_MATRIXS = 2  # /!\ must be > 1
            # Extract tags
            tag_product = tagg_type_node.get_tag_from_name(CONST.NODE_TYPE_PRODUCT)
            tag_sector = tagg_type_node.get_tag_from_name(CONST.NODE_TYPE_SECTOR)
            tag_exchange = tagg_type_node.get_tag_from_name(CONST.NODE_TYPE_EXCHANGE)
            # Extract nodes from tags
            nodes_tagged_as_products = tag_product.references
            nodes_tagged_as_sectors = tag_sector.references + tag_exchange.references
            # Use nodes entries to sort nodes from their parenthood relations and levels
            nodes_entries_tagged_as_products = []
            nodes_entries_tagged_as_products__levels = []
            nodes_entries_tagged_as_sectors = []
            nodes_entries_tagged_as_sectors__levels = []
            for (node, node_level) in zip(nodes_entries, nodes_entries__levels):
                if node in nodes_tagged_as_products:
                    nodes_entries_tagged_as_products.append(node)
                    nodes_entries_tagged_as_products__levels.append(node_level)
                if node in nodes_tagged_as_sectors:
                    nodes_entries_tagged_as_sectors.append(node)
                    nodes_entries_tagged_as_sectors__levels.append(node_level)
            # Create the two matrixs
            # 1 : sectors -> products
            # 2 : products -> sectors
            matrix_1 = _createMatrixFromFlux(
                nodes_entries_tagged_as_sectors,
                nodes_entries_tagged_as_products,
                transpose=True)
            matrix_2 = _createMatrixFromFlux(
                nodes_entries_tagged_as_products,
                nodes_entries_tagged_as_sectors)
            # Fuse the two matrixs
            # Header and indexs
            header = [_.name for _ in nodes_entries_tagged_as_sectors]
            index = [_.name for _ in nodes_entries_tagged_as_products]
            # Leave a white line between the two matrixs
            matrix = \
                matrix_1 + \
                [[None]*len(nodes_entries_tagged_as_sectors)]*(NB_ROWS_BETWEEN_MATRIXS - 1) + \
                [header] + \
                matrix_2
            # Panda table
            table = pd.DataFrame(
                matrix,
                index=(index + [None]*NB_ROWS_BETWEEN_MATRIXS + index),  # White lines between matrixs
                columns=header)
            # Save in sheets table and parameters
            sheets[CONST.TER_SHEET] = {}
            sheets[CONST.TER_SHEET].update(copy.deepcopy(SHEET_BY_DEFAULT))
            sheets[CONST.TER_SHEET]['table'] = table
            sheets[CONST.TER_SHEET]['name'] = self.xl_user_converter.get_user_sheet_name(CONST.TER_SHEET)
            sheets[CONST.TER_SHEET]['color'] = SHEET_MAIN_COLOR
            sheets[CONST.TER_SHEET]['write_index'] = True
            # Initialize default values for formating
            sheets[CONST.TER_SHEET].update(copy.deepcopy(SHEET_FORMATING))
            # For index and header, the height and width depend on nodes names
            max_size_header = len(max(header)) + 10
            max_size_index = len(max(index)) + 10
            sheets[CONST.TER_SHEET]['header']['default_height'] = max_size_header*10
            sheets[CONST.TER_SHEET]['index']['default_width'] = max_size_index*2
            # For content use fixed width and height
            sheets[CONST.TER_SHEET]['content']['compute_width'] = False
            sheets[CONST.TER_SHEET]['content']['default_width'] = 3
            sheets[CONST.TER_SHEET]['content']['default_height'] = 15
            # ----------------------------------------------------
            # Special formating - Colors for levels
            # Only if we have more than one level
            if self._max_nodes_level <= 1:
                return
            # Create a color gradient (from white to main_color) to fill nodes
            # cells depending on their respective level
            level_palette = sns \
                .color_palette(
                    "blend:#ffffff,#{}".format(SHEET_MAIN_COLOR),
                    self._max_nodes_level+1) \
                .as_hex()
            # Header - Get colors based on nodes levels
            colors__nodes_indexs = {}
            loop_iterator = enumerate(zip(
                nodes_entries_tagged_as_sectors,
                nodes_entries_tagged_as_sectors__levels))
            for index_node, (node, node_level) in loop_iterator:
                # Color for node level
                # /!\ Levels starts from 1, but the level palette starts from 0
                color = level_palette[node_level-1].replace('#', '')
                # Update Dict
                if not (color in colors__nodes_indexs.keys()):
                    colors__nodes_indexs[color] = []
                colors__nodes_indexs[color].append(index_node)
            # Header - Save special formating (colors based on nodes levels)
            second_matrix_starting_row = len(index) + (NB_ROWS_BETWEEN_MATRIXS - 1)
            for (color, nodes_indexs) in colors__nodes_indexs.items():
                # Convert as tuple to be used as dict key
                # Shift one col to the right, because we have the index col in first col
                cols = tuple(_+1 for _ in nodes_indexs)
                # Special formating for first matrix Header
                sheets[CONST.TER_SHEET]['spe_header'][cols] = \
                    copy.deepcopy(sheets[CONST.TER_SHEET]['header'])  # Keep other header's formating attributes
                sheets[CONST.TER_SHEET]['spe_header'][cols]['fill'] = \
                    PatternFill('solid', fgColor=color)
                # Special formating for second matrix Header
                sheets[CONST.TER_SHEET]['spe_content'][cols] = {}  # Init for
                sheets[CONST.TER_SHEET]['spe_content'][cols][(second_matrix_starting_row,)] = \
                    sheets[CONST.TER_SHEET]['spe_header'][cols]  # Copy of first matrix's header's formating attributes
            # Upper left corner of second matrix
            sheets[CONST.TER_SHEET]['spe_content'][(0,)] = {}
            sheets[CONST.TER_SHEET]['spe_content'][(0,)][(second_matrix_starting_row,)] = \
                sheets[CONST.TER_SHEET]['no_header']
            # Index - Get colors based on nodes levels
            colors__nodes_indexs = {}
            loop_iterator = enumerate(zip(
                nodes_entries_tagged_as_products,
                nodes_entries_tagged_as_products__levels))
            for index_node, (node, node_level) in loop_iterator:
                # Color for node level
                # /!\ Nodes levels start from 1, but the level_palette table starts from 0
                color = level_palette[node_level-1].replace('#', '')
                # Update Dict
                if not (color in colors__nodes_indexs.keys()):
                    colors__nodes_indexs[color] = []
                colors__nodes_indexs[color].append(index_node)
                colors__nodes_indexs[color].append(
                    index_node + len(nodes_entries_tagged_as_products) + NB_ROWS_BETWEEN_MATRIXS)
            # Index - Save special formating (colors based on nodes levels)
            sheets[CONST.TER_SHEET]['spe_content'][(0,)] = {}
            for (color, nodes_indexs) in colors__nodes_indexs.items():
                # Convert as tuple to be used as dict key
                row_id = tuple(nodes_indexs)
                # Special formating for Indexs
                sheets[CONST.TER_SHEET]['spe_content'][(0,)][row_id] = \
                    copy.deepcopy(sheets[CONST.TER_SHEET]['index'])
                sheets[CONST.TER_SHEET]['spe_content'][(0,)][row_id]['fill'] = \
                    PatternFill('solid', fgColor=color)

    def write_data_sheets(
        self,
        nodes_entries: list,
        sheets: dict
    ):
        """
        Rewrite all datas realted sheets their respective attributes and infos
        in one or some excel sheets. That includes:
        - Datas
        - Min Max
        - Constraints
        - Results
        - Analysis

        Parameters
        ----------
        :param nodes_entries: List of nodes sorted as they appear in nodes tables
        :type nodes_entries: list

        :param sheets: Contains the excel sheets
        :type sheets: dict (output, modified)
        """
        # ----------------------------------------------------
        if not self.has_at_least_one_data():
            return
        # ----------------------------------------------------
        # Sheet color
        SHEET_MAIN_DATA_COLOR = '8064A2'  # Green
        SHEET_MAIN_RESULTS_COLOR = '8064A2'  # Violet
        # Sheet formating infos
        SHEET_FORMATING_FOR_DATA = copy.deepcopy(SHEET_FORMATING_BY_DEFAULT)
        SHEET_FORMATING_FOR_DATA['header']['fill'] = PatternFill(
            'solid', fgColor=SHEET_MAIN_DATA_COLOR)
        SHEET_FORMATING_FOR_RESULTS = copy.deepcopy(SHEET_FORMATING_BY_DEFAULT)
        SHEET_FORMATING_FOR_RESULTS['header']['fill'] = PatternFill(
            'solid', fgColor=SHEET_MAIN_RESULTS_COLOR)
        # ----------------------------------------------------
        # Create tables
        table_data, table_min_max, table_constraints, table_results, table_analysis = \
            self._create_all_data_and_result_tables(
                default_header=False,
                reorder_tables=True,
                nodes_entries=nodes_entries)
        # ----------------------------------------------------
        # DATA_SHEET : Update excel sheet attributes
        sheets[CONST.DATA_SHEET] = copy.deepcopy(SHEET_BY_DEFAULT)
        sheets[CONST.DATA_SHEET]['name'] = \
            self.xl_user_converter.get_user_sheet_name(CONST.DATA_SHEET)
        sheets[CONST.DATA_SHEET]['color'] = SHEET_MAIN_DATA_COLOR
        sheets[CONST.DATA_SHEET]['table'] = table_data
        sheets[CONST.DATA_SHEET].update(copy.deepcopy(SHEET_FORMATING_FOR_DATA))
        # MIN_MAX_SHEET : Update excel sheet attributes
        if not table_min_max.empty:
            sheets[CONST.MIN_MAX_SHEET] = copy.deepcopy(SHEET_BY_DEFAULT)
            sheets[CONST.MIN_MAX_SHEET]['name'] = \
                self.xl_user_converter.get_user_sheet_name(CONST.MIN_MAX_SHEET)
            sheets[CONST.MIN_MAX_SHEET]['color'] = SHEET_MAIN_DATA_COLOR
            sheets[CONST.MIN_MAX_SHEET]['table'] = table_min_max
            sheets[CONST.MIN_MAX_SHEET].update(copy.deepcopy(SHEET_FORMATING_FOR_DATA))
        # CONSTRAINTS_SHEET : Update excel sheet attributes
        if not table_constraints.empty:
            sheets[CONST.CONSTRAINTS_SHEET] = copy.deepcopy(SHEET_BY_DEFAULT)
            sheets[CONST.CONSTRAINTS_SHEET]['name'] = \
                self.xl_user_converter.get_user_sheet_name(CONST.CONSTRAINTS_SHEET)
            sheets[CONST.CONSTRAINTS_SHEET]['color'] = SHEET_MAIN_DATA_COLOR
            sheets[CONST.CONSTRAINTS_SHEET]['table'] = table_constraints
            sheets[CONST.CONSTRAINTS_SHEET].update(copy.deepcopy(SHEET_FORMATING_FOR_DATA))
            # Demarquation // id
            col_name = self.xl_user_converter.get_user_col_name(
                CONST.CONSTRAINTS_SHEET, CONST.CONSTRAINT_ID)
            cols = tuple(range(len(table_constraints.columns)))
            prev_id = None
            rows = []
            for _, id in enumerate(table_constraints[col_name]):
                row = _ - 1
                if prev_id is None:
                    prev_id = id
                else:
                    if prev_id == id:
                        rows.append(row)
                    else:
                        prev_id = id
            if len(rows) > 0:
                rows = tuple(rows)
                sheets[CONST.CONSTRAINTS_SHEET]['spe_content'][cols] = {}
                sheets[CONST.CONSTRAINTS_SHEET]['spe_content'][cols][rows] = \
                    copy.deepcopy(sheets[CONST.CONSTRAINTS_SHEET]['content'])
                sheets[CONST.CONSTRAINTS_SHEET]['spe_content'][cols][rows]['border'] = \
                    Border(
                        left=Side(border_style="thin", color=COLOR_BLACK),
                        right=Side(border_style="thin", color=COLOR_BLACK),
                        top=Side(border_style="none"))
        # ----------------------------------------------------
        # RESULTS_SHEET : Update excel sheet attributes
        if not table_results.empty:
            sheets[CONST.RESULTS_SHEET] = copy.deepcopy(SHEET_BY_DEFAULT)
            sheets[CONST.RESULTS_SHEET]['name'] = \
                self.xl_user_converter.get_user_sheet_name(CONST.RESULTS_SHEET)
            sheets[CONST.RESULTS_SHEET]['color'] = SHEET_MAIN_RESULTS_COLOR
            sheets[CONST.RESULTS_SHEET]['table'] = table_results
            sheets[CONST.RESULTS_SHEET].update(SHEET_FORMATING_FOR_RESULTS)
        # ----------------------------------------------------
        # ANALYSIS_SHEET : Update excel sheet attributes
        if not table_analysis.empty:
            sheets[CONST.ANALYSIS_SHEET] = copy.deepcopy(SHEET_BY_DEFAULT)
            sheets[CONST.ANALYSIS_SHEET]['name'] = \
                self.xl_user_converter.get_user_sheet_name(CONST.ANALYSIS_SHEET)
            sheets[CONST.ANALYSIS_SHEET]['color'] = SHEET_MAIN_RESULTS_COLOR
            sheets[CONST.ANALYSIS_SHEET]['table'] = table_analysis
            sheets[CONST.ANALYSIS_SHEET].update(SHEET_FORMATING_FOR_RESULTS)

    def get_as_dict(self):
        """
        """
        # Init output structure
        output_dict = {}
        # Parse tags
        output_dict["taggs"] = {}
        for tag_type in self.taggs.keys():
            output_dict["taggs"][tag_type] = {}
            tagg_names_sorted = sorted(self.taggs[tag_type].keys())
            for tagg_name in tagg_names_sorted:
                tagg = self.taggs[tag_type][tagg_name]
                output_dict["taggs"][tag_type][tagg_name] = tagg.get_as_dict()
        # Parse nodes
        nodes_names_sorted = sorted(self.nodes.keys())
        output_dict["nodes"] = {}
        for node_name in nodes_names_sorted:
            node = self.nodes[node_name]
            output_dict["nodes"][node_name] = node.get_as_dict()
        # Parse flux
        flux_names_sorted = sorted(self.flux.keys())
        output_dict["flux"] = {}
        for flux_name in flux_names_sorted:
            flux = self.flux[flux_name]
            output_dict["flux"][flux_name] = flux.get_as_dict()
        # Parse constraintes
        constraints_ids_sorted = sorted(self.constraints.keys())
        output_dict["constraints"] = {}
        for constraint_id in constraints_ids_sorted:
            constraint = self.constraints[constraint_id]
            output_dict["constraints"][constraint_id] = [_.get_as_dict() for _ in constraint]
        # End
        return output_dict

    def update_mfa_dict(self, mfa_dict: dict):
        """
        _summary_

        Parameters
        ----------
        :param mfa_dict: Dictionnaire des données
        :type mfa_dict: dict
        """
        self.update_mfa_dict_taggs(mfa_dict)
        self.update_mfa_dict_nodes(mfa_dict)
        self.update_mfa_dict_flux(mfa_dict)
        self.update_mfa_dict_data_and_result(mfa_dict)

    def update_mfa_dict_taggs(self, mfa_dict: dict):
        """
        _summary_

        Parameters
        ----------
        :param mfa_dict: Dictionnaire des données
        :type mfa_dict: dict
        """
        # Check if we have tags to save
        if not self.has_at_least_one_tagg():
            return
        # Specify columns for table flux
        taggs_extra_infos_names = self.taggs_extra_infos_names
        table_taggs__columns = CONST.TAG_SHEET_COLS + taggs_extra_infos_names
        # Fill table tag with types in specific order
        table_taggs = []
        for tagg_type in [CONST.TAG_TYPE_LEVEL, CONST.TAG_TYPE_NODE, CONST.TAG_TYPE_DATA, CONST.TAG_TYPE_FLUX]:
            antagonists_checked = []
            for tagg in self.taggs[tagg_type].values():
                # Already taken in account as antagonist tagg ?
                if tagg in antagonists_checked:
                    continue
                # Tag groups infos
                name = tagg.name_unformatted
                tags = tagg.tags_str
                # Specific case with antagonist
                if tagg.has_antagonists():
                    for antagonist_tagg in tagg.antagonists_taggs:
                        name += '/' + antagonist_tagg.name_unformatted
                        tags += '/' + antagonist_tagg.tags_str
                        antagonists_checked.append(antagonist_tagg)
                # Create table line with corresponding data
                line_tagg = [
                    name,
                    tagg_type,
                    tags,
                    tagg.is_palette,
                    tagg.colormap,
                    tagg.colors]
                # Add extra info cols if needed
                for extra_info_name in taggs_extra_infos_names:
                    if extra_info_name in tagg.extra_infos.keys():
                        line_tagg.append(tagg.extra_infos[extra_info_name])
                    else:
                        line_tagg.append(None)
                # We can add it directly in the table
                table_taggs.append(line_tagg)
        table_taggs = pd.DataFrame(table_taggs, columns=table_taggs__columns)
        # Drop column that have no values
        table_taggs.dropna(axis=1, how='all', inplace=True)
        # Cast NaN as None because if you have None in a float column,
        # panda transform it as NaN -> cant compare tests after
        table_taggs.replace({np.nan: None}, inplace=True)
        # Update MFA dict
        mfa_dict[CONST.TAG_SHEET] = table_taggs

    def update_mfa_dict_nodes(self, mfa_dict: dict):
        """
        _summary_

        Parameters
        ----------
        :param mfa_dict: Dictionnaire des données
        :type mfa_dict: dict
        """
        # Columns for tags
        columns_taggs_names = [tagg.name_unformatted for tagg in self.node_taggs]
        columns_taggs_names += [tagg.name_unformatted for tagg in self.level_taggs]
        # If we have node type tag (product:sector:exchange),
        # then it must be the first column of the tags columns
        if CONST.NODE_TYPE in columns_taggs_names:
            columns_taggs_names.remove(CONST.NODE_TYPE)
            columns_taggs_names.insert(0, CONST.NODE_TYPE)
        # Specify columns for node table
        nodes_extra_infos_names = self.nodes_extra_infos_names
        table_node__columns = \
            [CONST.NODES_LEVEL, CONST.NODES_NODE, CONST.NODES_MAT_BALANCE, CONST.NODES_COLOR] + \
            columns_taggs_names + \
            [CONST.NODES_DEFINITIONS] + \
            nodes_extra_infos_names
        # Fill table node
        table_node = []
        nodes_processed = []
        for node in self.nodes.values():
            if node not in nodes_processed:
                self._create_node_line(
                    node,
                    node.level,
                    columns_taggs_names,
                    nodes_extra_infos_names,
                    table_node,
                    nodes_processed)
        table_node = pd.DataFrame(table_node, columns=table_node__columns)
        # Drop column that have no values
        table_node.dropna(axis=1, how='all', inplace=True)
        # Cast NaN as None because if you have None in a float column,
        # panda transform it as NaN -> cant compare tests after
        table_node.replace({np.nan: None}, inplace=True)
        # Update MFA dict
        mfa_dict[CONST.NODES_SHEET] = table_node

    def _create_node_line(
        self,
        node: Node,
        node_level: int,
        columns_taggs_names: list,
        nodes_extra_infos_names: list,
        table_node: list,
        nodes_processed: list,
        process_children: bool = True
    ):
        """
        _summary_

        Parameters
        ----------
        TODO
        """
        # Create table line with corresponding data
        line_node = [
            node_level,
            node.name,
            node.mat_balance,
            node.color]
        # Add tags
        line_node += node.get_tags_from_taggroups(
            columns_taggs_names, return_names_instead_of_refs=True)
        # Add definition
        line_node.append(node.definition)
        # Add extra info cols if needed
        for extra_info_name in nodes_extra_infos_names:
            if extra_info_name in node.extra_infos.keys():
                line_node.append(node.extra_infos[extra_info_name])
            else:
                line_node.append(None)
        # Add line to the table
        table_node.append(line_node)
        # If we have children for this node, we add them directly under
        if node.has_at_least_one_child() and process_children:
            for childrengrp_id, childgroup in enumerate(node.children_grps):
                # Do we need to add a new line for current node : ie multiple childgroup
                if childrengrp_id > 0:
                    table_node.append(line_node)
                # Recursivly process children
                for child in childgroup:
                    self._create_node_line(
                        child,
                        node_level+1,
                        columns_taggs_names,
                        nodes_extra_infos_names,
                        table_node,
                        nodes_processed,
                        process_children=(child not in nodes_processed))
        # Ok node processed
        nodes_processed.append(node)

    def update_mfa_dict_flux(self, mfa_dict: dict):
        """
        _summary_

        Parameters
        ----------
        :param mfa_dict: Dictionnaire des données
        :type mfa_dict: dict
        """
        # Check if we have flux to save
        if not self.has_at_least_one_flux():
            return
        # Init table flux
        table_flux = []
        # Fill table flux
        for flux in self.flux.values():
            # Create table line with corresponding data
            line_flux = [
                flux.orig.name,
                flux.dest.name]
            # We can add it directly in the table
            table_flux.append(line_flux)
        # Cast to dataFrame
        table_flux = pd.DataFrame(table_flux, columns=CONST.FLUX_SHEET_COLS)
        # Update MFA dict
        mfa_dict[CONST.FLUX_SHEET] = table_flux

    def update_mfa_dict_data_and_result(
        self,
        mfa_dict: dict
    ):
        """
        _summary_

        Parameters
        ----------
        :param mfa_dict: Dictionnaire des données
        :type mfa_dict: dict
        """
        # Check if we have data to save
        # TODO : probably need to check also for results data
        # if not self.has_at_least_one_data():
        #     return
        # Create tables
        table_data, table_min_max, table_constraints, table_results, table_analysis = \
            self._create_all_data_and_result_tables()
        # Update MFA dict
        if not table_data.empty:
            mfa_dict[CONST.DATA_SHEET] = table_data
        if not table_min_max.empty:
            mfa_dict[CONST.MIN_MAX_SHEET] = table_min_max
        if not table_constraints.empty:
            mfa_dict[CONST.CONSTRAINTS_SHEET] = table_constraints
        if not table_results.empty:
            mfa_dict[CONST.RESULTS_SHEET] = table_results
        if not table_analysis.empty:
            mfa_dict[CONST.ANALYSIS_SHEET] = table_analysis

    def _create_all_data_and_result_tables(
        self,
        default_header=True,
        reorder_tables=False,
        nodes_entries=None
    ):
        """
        _summary_

        Parameters
        ----------
        :param default_header: Use default header or not
        :type default_header: bool (default=False)

        :param nodes_entries: Node ordering to follow when writing tables. If None, not order to follow.
        :type nodes_entries: None | list (default=None)
        """
        # Columns for tags
        columns_datataggs_names = [tagg.name_unformatted for tagg in self.data_taggs]
        columns_fluxtaggs_names = [tagg.name_unformatted for tagg in self.flux_taggs]
        # ----------------------------------------------------
        # Specify all columns for data table
        data_extra_infos_names = []  # = self.data_extra_infos_names
        table_data__cols = \
            CONST.DATA_SHEET_COLS_1 + \
            columns_datataggs_names + \
            columns_fluxtaggs_names + \
            CONST.DATA_SHEET_COLS_2 + \
            data_extra_infos_names
        if not default_header:
            table_data__cols = \
                [self.xl_user_converter.get_user_col_name(CONST.DATA_SHEET, _) for _ in table_data__cols]
        # Specify all columns for min_max table
        table_min_max__cols = \
            CONST.MIN_MAX_SHEET_COLS_1 + \
            columns_datataggs_names + \
            columns_fluxtaggs_names + \
            CONST.MIN_MAX_SHEET_COLS_2
        if not default_header:
            table_min_max__cols = \
                [self.xl_user_converter.get_user_col_name(CONST.MIN_MAX_SHEET, _) for _ in table_min_max__cols]
        # Specify all columns for constraints table
        table_constraints__cols = \
            CONST.CONSTRAINT_SHEET_COLS_1 + \
            columns_datataggs_names + \
            columns_fluxtaggs_names + \
            CONST.CONSTRAINT_SHEET_COLS_2
        if not default_header:
            table_constraints__cols = \
                [self.xl_user_converter.get_user_col_name(CONST.CONSTRAINTS_SHEET, _) for _ in table_constraints__cols]
        # ----------------------------------------------------
        # Specify all columns for result table
        table_results__cols = \
            CONST.RESULTS_SHEET_COLS_1 + \
            columns_datataggs_names + \
            columns_fluxtaggs_names + \
            CONST.RESULTS_SHEET_COLS_2
        if not default_header:
            table_results__cols = \
                [self.xl_user_converter.get_user_col_name(CONST.RESULTS_SHEET, _) for _ in table_results__cols]
        # Specify all columns for analysis table
        table_analysis__cols = \
            CONST.ANALYSIS_SHEET_COLS_1 + \
            columns_datataggs_names + \
            columns_fluxtaggs_names + \
            CONST.ANALYSIS_SHEET_COLS_2
        # ----------------------------------------------------
        # Init empty tables
        table_data = []
        table_min_max = []
        table_constraints = []
        table_results = []
        table_analysis = []
        # ----------------------------------------------------
        # Write data sheets
        # Fill table data : Loop on flux, because data are related to flux
        if nodes_entries is None:
            nodes_entries = self.nodes.values()
        # Keep in memory the list of already processed flux
        flux_processed = []
        for node in nodes_entries:
            for flux in node.output_flux:
                # Check if flux has not been already processed
                if flux in flux_processed:
                    continue
                # ----------------------------------------------------
                # Fill tables from data related informations
                for data in flux.datas:
                    # Update data sheet
                    data.update_table(
                        columns_datataggs_names,
                        columns_fluxtaggs_names,
                        data_extra_infos_names,
                        table_data)
                    # Update min_max sheet
                    data.min_max.update_table(
                        columns_datataggs_names,
                        columns_fluxtaggs_names,
                        [],  # No extra info for min_max
                        table_min_max)
                    # Update constraints sheets
                    for constraints in data.constraints.values():
                        for constraint in constraints:
                            constraint.update_table(
                                columns_datataggs_names,
                                columns_fluxtaggs_names,
                                [],  # No extra info for constraints
                                table_constraints)
                # ----------------------------------------------------
                # Fill remaining data info that are related to flux
                # Update min_max sheet
                flux.min_max.update_table(
                    columns_datataggs_names,
                    columns_fluxtaggs_names,
                    [],  # No extra info for min_max
                    table_min_max)
                # Update constraints sheets
                for constraints in flux.constraints.values():
                    for constraint in constraints:
                        constraint.update_table(
                            columns_datataggs_names,
                            columns_fluxtaggs_names,
                            [],  # No extra info for constraints
                            table_constraints)
                # ----------------------------------------------------
                # Fill tables from results related informations
                for result in flux.results:
                    # Update result sheet
                    result.update_table(
                        columns_datataggs_names,
                        columns_fluxtaggs_names,
                        data_extra_infos_names,
                        table_results,
                        as_result=True,
                        table_for_analysis=table_analysis)
                # ----------------------------------------------------
                # Keep track of processed flux
                flux_processed.append(flux)
        # ----------------------------------------------------
        # Create panda tables with correct headers
        table_data = pd.DataFrame(
            table_data,
            columns=table_data__cols)
        table_min_max = pd.DataFrame(
            table_min_max,
            columns=table_min_max__cols)
        table_constraints = pd.DataFrame(
            table_constraints,
            columns=table_constraints__cols)
        table_results = pd.DataFrame(
            table_results,
            columns=table_results__cols)
        # Special case for analysis sheet
        try:
            if len(table_analysis[0]) > len(table_analysis__cols):
                table_analysis__cols.append('Ai constraints ids')
        except Exception:
            pass
        table_analysis = pd.DataFrame(
            table_analysis,
            columns=table_analysis__cols)
        # ----------------------------------------------------
        # Contraints table must be sorted by ids
        id_col = CONST.CONSTRAINT_ID
        if not default_header:
            id_col = self.xl_user_converter.get_user_col_name(
                CONST.CONSTRAINTS_SHEET, CONST.CONSTRAINT_ID)
        table_constraints.sort_values(
            id_col,
            axis=0,
            ascending=True,
            inplace=True)
        # ----------------------------------------------------
        # Sort all table accordingly to tags
        if reorder_tables:
            tables_reordered = []
            # Ordering priority for col header / data taggroups
            ordering_priority_for_datataggs = [
                tagg.name_unformatted
                for tagg in reversed(self.data_taggs)]
            # Ordering priority for content for each col header / data tags
            ordering_priority_for_datatags = [
                [tag.name_unformatted for tag in tagg.tags.values()]
                for tagg in reversed(self.data_taggs)]
            for table in (table_data, table_min_max, table_constraints, table_results, table_analysis):
                # Ordering table
                tables_reordered.append(
                    _reorderTable(
                        table,
                        ordering_priority_for_datataggs.copy(),
                        ordering_priority_for_datatags.copy()))
            table_data, table_min_max, table_constraints, table_results, table_analysis = tables_reordered
        # ----------------------------------------------------
        # Drop column that have no values
        table_data.dropna(axis=1, how='all', inplace=True)
        table_min_max.dropna(axis=1, how='all', inplace=True)
        table_constraints.dropna(axis=1, how='all', inplace=True)
        table_results.dropna(axis=1, how='all', inplace=True)
        table_analysis.dropna(axis=1, how='all', inplace=True)
        # Cast NaN as None because if you have None in a float column,
        # panda transform it as NaN -> cant compare tests after
        table_data.replace({np.nan: None}, inplace=True)
        table_min_max.replace({np.nan: None}, inplace=True)
        table_constraints.replace({np.nan: None}, inplace=True)
        table_results.replace({np.nan: None}, inplace=True)
        table_analysis.replace({np.nan: None}, inplace=True)
        # Outputs
        return table_data, table_min_max, table_constraints, table_results, table_analysis

    def autocompute_nodes_types(self):
        """
        Compute all nodes types. Nodes' types can be :
            - SR = Single-Root : Node that have no child and no parent
            - PR = Parent-Root : Node that have at least one child but no parent
            - PC = Parent-Child : Node that have at least one child and at least one parent
            - BC = Base-Child : Node that have no child and at least one parent
        """
        # Loop on all nodes
        for node in self.nodes.values():
            # Avoid numerous calls to parents / children existence checks
            has_parents = node.has_parents()
            has_at_least_one_child = node.has_at_least_one_child()
            # Build nodetype dict nodetypes are
            # 'SR': Single root, 'PR': parent root, 'BC': base child, 'PC': parent child
            if has_parents and has_at_least_one_child:
                node.type = 'PC'
            elif has_parents:
                node.type = 'BC'
            elif has_at_least_one_child:
                node.type = 'PR'
            else:
                node.type = 'SR'
            # Special cases - importation nette
            if ' nette' in node.name:
                node.type = 'PR'
