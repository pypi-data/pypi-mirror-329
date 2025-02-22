from matplotlib import pyplot as plt
import numpy as np
from scipy import stats
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from zipfile import BadZipFile

# Main FGMK Plotter
def fgmk3(figure_number, fig_title, config, *datasets):
    # Config
    pltshow = config.get("pltshow", 1)
    pltlegend = config.get("pltlegend", 0)
    grid = config.get("grid", 1)
    ptype = config.get("ptype", "plot")
    figt = config.get("figt", "")
    deltxt = config.get("deltxt", 0)
    dim = config.get("dim", (10, 6))
    splots = config.get("splots", 1)
    lbl = config.get("lbl", [])
    lst = config.get("lst", [])
    mrk = config.get("mrk", [])
    lc = config.get("lc", [])
    while len(lc) < len(lbl):
        lc.append("#{:06x}".format(np.random.randint(0, 0xFFFFFF)))
    loc_ = config.get('loc', 'best')

    # General
    bottom_text = r'$\bf{Figure\ }$' + str(figure_number) + r': ' + str(figt)


    # Display
    if splots == 1:
        plt.figure(figure_number, figsize=dim)
        for i, dataset in enumerate(datasets):
            x_data, y_data, xlabel, ylabel = dataset
            if ptype == "stem":
                plt.plot(x_data, y_data, basefmt=" ", linefmt='-', markerfmt='o',
                        label=lbl[i] if lbl else None,
                        linestyle=lst[i] if lst else None,
                        marker=mrk[i] if mrk else None,
                        color=lc[i] if lc else None)
            elif ptype == "hist":
                plt.plot(x_data, bins=y_data,
                        label=lbl[i] if lbl else None,
                        linestyle=lst[i] if lst else None,
                        marker=mrk[i] if mrk else None,
                        color=lc[i] if lc else None)
            else:
                plt.plot(x_data, y_data,
                        label=lbl[i] if lbl else None,
                        linestyle=lst[i] if lst else None,
                        marker=mrk[i] if mrk else None,
                        color=lc[i] if lc else None)
        plt.xlabel(xlabel)
        plt.ylabel(ylabel)

        # Layout Adjustments
        plt.title(fig_title)
        plt.grid(grid)
        if figt:
            plt.text(0.5, -0.1 * (splots) + deltxt, bottom_text,transform=plt.gca().transAxes, horizontalalignment='center', verticalalignment='center', fontsize=10)
        plt.tight_layout()
        if pltlegend ==1:
            plt.legend(loc=loc_)
        if pltshow == 1:
            plt.show()
    else:
        fig_setup, axs = plt.subplots(splots, 1, figsize=dim)
        if isinstance(fig_title, str):
            axs[0].set_title(fig_title)

        for i, dataset in enumerate(datasets):
            x_data, y_data, xlabel, ylabel = dataset
            ax = axs[i]
            if ptype == "stem":
                ax.plot(x_data, y_data, basefmt=" ", linefmt='-', markerfmt='o',
                        label=lbl[i] if lbl else None,
                        linestyle=lst[i] if lst else None,
                        marker=mrk[i] if mrk else None,
                        color=lc[i] if lc else None)
            elif ptype == "hist":
                ax.plot(x_data, bins=y_data,
                        label=lbl[i] if lbl else None,
                        linestyle=lst[i] if lst else None,
                        marker=mrk[i] if mrk else None,
                        color=lc[i] if lc else None)
            else:
                ax.plot(x_data, y_data,
                        label=lbl[i] if lbl else None,
                        linestyle=lst[i] if lst else None,
                        marker=mrk[i] if mrk else None,
                        color=lc[i] if lc else None)
            ax.set_xlabel(xlabel)
            ax.set_ylabel(ylabel)

            # Adjust layout and display
            if not isinstance(fig_title, str):
                ax.set_title(fig_title)
            ax.grid(grid)
            if pltlegend ==1:
                ax.legend(loc=loc_)

        if figt:
            plt.text(0.5, -0.1 * (splots) + deltxt, bottom_text,transform=plt.gca().transAxes, horizontalalignment='center', verticalalignment='center', fontsize=10)
        plt.tight_layout()
        if pltshow == 1:
            plt.show()


# Saving Data to Excel Sheets
def toSheet(workbook_name, sheetname, **kwargs):
    data_dict = {}
    max_length = 0

    # Generate Dictionary
    for variable_name, values in kwargs.items():
        if not isinstance(values, list):  # list or var
            values = [values]
        if len(values) > max_length:  # Store max length array
            max_length = len(values)
        data_dict[variable_name] = values  # Populate

    # Equal length normalization excel nonsense
    for key in data_dict:
        if len(data_dict[key]) < max_length:
            data_dict[key].extend([None] * (max_length - len(data_dict[key])))

    # Logistics
    df = pd.DataFrame(data_dict)
    current_dir = os.getcwd()
    workbook_path = os.path.join(current_dir, workbook_name)

    # Workbook Sanity Checks
    if os.path.exists(workbook_path):
        try:
            book = load_workbook(workbook_path)
        except (FileNotFoundError, KeyError, BadZipFile):
            book = Workbook()
            book.remove(book.active)  # Remove default sheet
    else:
        book = Workbook()
        book.remove(book.active)  # Remove default sheet

    if sheetname not in book.sheetnames:
        book.create_sheet(title=sheetname)

    # Save the Wb
    book.save(workbook_path)

    # Process
    with pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheetname, index=False)

# Help
def fgmk_help():
    print("fgmk3:")
    print("    Mandatory:")
    print("      Figure Number, Title(s), Config or {}, data[x0, y0, xlabel, ylabel]")
    print("    Optional:")
    print("      x: x values for data sets")
    print("      y: y values for data sets")
    print("      lbl: Label for data sets")
    print("      lst: Line Style for data sets")
    print("      mrk: Line Marker for data sets")
    print("      lc: Color for data sets")
    print("      figt: Figure text")
    print("      deltxt: Figure text offset")
    print("      grid: Enable grid (default: 1)")
    print("      splots: Number of subplots")
    print("      ptype: Type of Plot")
    print("      dim: 1x2 Tuple for figure dimensions\n")

    print("toSheet:")
    print("    Mandatory:")
    print("      Excel Name, Sheet Name")
    print("    Optional:")
    print("      Any array or numeric value in dictionary form, such as Var=[21] or Var=Var\n")

    print("fgmk, fgmk2, and hgmk2 are deprecated as of 2024-11-27")

#################################################################################################################

# Legacy FGMK2s

#   Single Input
def fgmk(n,x,y,xlabel,ylabel, titlestr, figtext = None, glabel = None, grid = None,dim=None):
    print("Deprecated, use FGMK3")
#   Histogram
def hgmk2(n, data, xlabel, ylabel, t, l=None, lst='-', c='tab:blue', fig=None, grid=1, dim=None,
            conf_intervals=None, mark_values=None, mark_labels=None, average=False, bell_curve=False,
            xmax=None):
    print("Deprecated, use FGMK3")
#   Multiple Inputs
def fgmk2(n, x0, y0, xlabel, ylabel, t, *args, splots=1, lst='-', c='tab:blue', l=" ", basefmt=" ", grid=1, dim=(10,6), fig=None, fig_offset=0,
          mark_values=None, mark_labels=None, ptype="plot", conf_intervals=None, average=False, bell_curve=False, xmax=None,
          pltshow=1, pltlegend=1,
          **kwargs):
    # Figure Setup
    fig_setup, axs = plt.subplots(splots, 1, figsize=dim)

    # If only one subplot, axs -> list
    if splots == 1:
        axs = [axs]

    # Bell Curve Fit
    def plot_bell_curve(ax, data):
        mu, std = stats.norm.fit(data)
        xmin, xmax = ax.get_xlim()
        x = np.linspace(xmin, xmax, 100)
        p = stats.norm.pdf(x, mu, std)
        ax.plot(x, p * len(data) * np.diff(ax.get_xticks())[0], 'k--', linewidth=2)

    # Graph Handler Dataset 0
    lst_first = kwargs.get('lst1', lst)
    c_first = kwargs.get('c1', c)
    l_first = kwargs.get('l1', l)

    if ptype == "stem":
        axs[0].stem(x0, y0, basefmt=basefmt, linefmt=lst_first, markerfmt=c_first, label=l_first)
    elif ptype == "hist":
        axs[0].hist(x0, bins=y0, color=c_first)
        if conf_intervals is not None:
            for ci in conf_intervals:
                axs[0].axvline(ci, color='r', linestyle='--', label=f'CI {ci}')
        if average:
            avg = np.mean(x0)
            axs[0].axvline(avg, color='g', linestyle='-', label=f'Average: {avg:.2f}')
        if bell_curve:
            plot_bell_curve(axs[0], x0)
        if xmax is not None:
            axs[0].set_xlim(right=xmax)
    else:
        axs[0].plot(x0, y0, linestyle=lst_first, color=c_first, label=l_first)

    # Main Labels, Dataset 0
    axs[0].set_title(t)
    axs[0].set_xlabel(xlabel)
    axs[0].set_ylabel(ylabel)
    if grid:
        axs[0].grid(True)

    # Additional Datasets Handler
    for i in range(1, splots):
        x_data = args[(i - 1) * 5]      # x1, x2, ..., xN
        y_data = args[(i - 1) * 5 + 1]  # y1, y2, ..., yN
        title = args[(i - 1) * 5 + 2]   # title1, title2, ..., titleN
        xlabel = args[(i - 1) * 5 + 3]  # xlabel1, xlabel2, ..., xlabelN
        ylabel = args[(i - 1) * 5 + 4]  # ylabel1, ylabel2, ..., ylabelN

        # Optional Line Style and Color Handler
        lst_key = f'lst{i+1}'
        c_key = f'c{i+1}'
        lst = kwargs.get(lst_key, '-')
        c = kwargs.get(c_key, 'tab:blue')

        if ptype == "stem":
            axs[i].stem(x_data, y_data, basefmt=basefmt, linefmt=lst, markerfmt=c)
        elif ptype == "hist":
            axs[i].hist(x_data, bins=y_data, color=c)
            if conf_intervals is not None:
                for ci in conf_intervals:
                    axs[i].axvline(ci, color='r', linestyle='--', label=f'CI {ci}')
            if average:
                avg = np.mean(x_data)
                axs[i].axvline(avg, color='g', linestyle='-', label=f'Average: {avg:.2f}')
            if bell_curve:
                plot_bell_curve(axs[i], x_data)
            if xmax is not None:
                axs[i].set_xlim(right=xmax)
        else:
            axs[i].plot(x_data, y_data, linestyle=lst, color=c)

        axs[i].set_title(title)
        axs[i].set_xlabel(xlabel)
        axs[i].set_ylabel(ylabel)
        if grid:
            axs[i].grid(True)

    # Marker Handler
    if mark_values is not None and mark_labels is not None:
        for ax in axs:
            ax.set_xticks(mark_values)
            ax.set_xticklabels(mark_labels)

    # Figure Name Handler
    if fig is not None:
        plt.text(0.5, -0.1 * (splots*1.15 + fig_offset), r'$\bf{Figure\ }$' + str(n) + r': ' + str(fig), transform=plt.gca().transAxes,
                 horizontalalignment='center', verticalalignment='center', fontsize=10)

    plt.tight_layout()
    if pltshow == 1:
        plt.show()
    if pltlegend == 1:
        plt.legend()
