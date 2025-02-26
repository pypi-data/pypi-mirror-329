from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from module_placeholder.templates import get_template_path


def create_workbook_from_template(output_file, template_name, dataframe):
    template_path = get_template_path(template_name)
    book = load_workbook(template_path)

    # create a new sheet
    sheet = book.create_sheet("data")

    # Load dataframe into new sheet
    for row in dataframe_to_rows(dataframe, index=False, header=True):
        sheet.append(row)

    # Save the modified excel at desired location
    book.save(output_file.name)
