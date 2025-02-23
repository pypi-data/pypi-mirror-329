import os
import argparse
import concurrent.futures
from PyPDF2 import PdfReader
from docx import Document
import openpyxl


def extract_text_from_file(file_path):
    """Extract text based on the file extension."""
    _, file_extension = os.path.splitext(file_path)
    file_extension = file_extension.lower()

    extractors = {
        '.txt': extract_text_from_txt,
        '.pdf': extract_text_from_pdf,
        '.docx': extract_text_from_docx,
        '.xlsx': extract_text_from_xlsx,
    }

    if file_extension in extractors:
        return extractors[file_extension](file_path)
    elif file_extension in ['.py', '.java', '.js', '.html', '.css', '.json', '.xml', '.c', '.cpp']:
        return extract_text_from_code(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_extension}")

def extract_text_from_txt(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()

def extract_text_from_pdf(file_path):
    text_parts = []
    with open(file_path, 'rb') as f:
        reader = PdfReader(f)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    return "\n".join(text_parts)

def extract_text_from_docx(file_path):
    doc = Document(file_path)
    text_parts = [paragraph.text for paragraph in doc.paragraphs]
    return "\n".join(text_parts)

def extract_text_from_xlsx(file_path):
    text_parts = []
    workbook = openpyxl.load_workbook(file_path)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            row_text = " ".join(str(cell) if cell is not None else '' for cell in row)
            text_parts.append(row_text)
    return "\n".join(text_parts)

def extract_text_from_code(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()

def process_single_file(file_path, args):
    """Process a single file and return the result."""
    try:
        if args.verbose:
            print(f"Processing file: {file_path}")
        
      
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        extracted_text = extract_text_from_file(file_path)
        
      
        if args.output:
            if os.path.isdir(args.output):
                base_name = os.path.basename(file_path)
                output_file = os.path.join(args.output, f"{os.path.splitext(base_name)[0]}_extracted.txt")
            else:
                if len(args.file_paths) > 1:
                    raise ValueError("When processing multiple files, --output must be a directory.")
                output_file = args.output
        else:
            output_file = f"{os.path.splitext(os.path.basename(file_path))[0]}_extracted.txt"
        
       
        if os.path.exists(output_file) and not args.force:
            overwrite = input(f"File {output_file} exists. Overwrite? (y/n): ").strip().lower()
            if overwrite != 'y':
                return (file_path, None, "Skipped due to user choice")
        
       
        with open(output_file, 'w', encoding='utf-8') as out_file:
            out_file.write(extracted_text)
        
        return (file_path, output_file, None)
    
    except Exception as e:
        return (file_path, None, str(e))

def main():
    parser = argparse.ArgumentParser(
        description="tfq0tool: is a command-line utility for extracting text from various file formats, including text files, PDFs, Word documents, spreadsheets, and code files in popular programming languages.",
        epilog="Examples:\n"
               "  tfq0tool sample.pdf\n"
               "  tfq0tool document.docx --output ./extracted\n"
               "  tfq0tool file1.txt file2.txt --threads 4 --verbose",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument("file_paths", nargs='*', help="Path(s) to the file(s) for text extraction")
    parser.add_argument("-o", "--output", type=str, help="Output directory or file (must be a directory if multiple files)")
    parser.add_argument("-t", "--threads", type=int, default=1, help="Number of threads for parallel processing (default: 1)")
    parser.add_argument("-v", "--verbose", action="store_true", help="Enable verbose output")
    parser.add_argument("-f", "--force", action="store_true", help="Overwrite files without prompting")
    
    args = parser.parse_args()


    if not args.file_paths:
        parser.print_help()
        return


    if args.output and not os.path.isdir(args.output) and len(args.file_paths) > 1:
        parser.error("When processing multiple files, --output must be a directory.")

 
    with concurrent.futures.ThreadPoolExecutor(max_workers=args.threads) as executor:
        futures = [executor.submit(process_single_file, fp, args) for fp in args.file_paths]
        for future in concurrent.futures.as_completed(futures):
            file_path, output_file, error = future.result()
            if error:
                print(f"Error processing {file_path}: {error}")
            elif output_file:
                print(f"Success: {file_path} -> {output_file}")

if __name__ == "__main__":
    main()