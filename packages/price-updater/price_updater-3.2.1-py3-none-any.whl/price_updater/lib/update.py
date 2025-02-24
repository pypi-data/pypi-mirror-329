from typing import Tuple, List, Any
from time import sleep
from importlib.resources import files
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from bs4 import BeautifulSoup
from requests import get, exceptions, Response

from price_updater.lib.settings import settings
from price_updater.lib.currency import currency, Currency
from price_updater.lib.asset import Asset


class Update:
    def __init__(self) -> None:
        self.api_status: str = settings.get_api_status()

    def get_logo_path(self, name: str) -> str:
        if "ishares" in name:
            with (
                files("price_updater.images.asset_logo")
                .joinpath("ishares.png")
                .open("rb") as ishares_logo
            ):
                return str(ishares_logo.name)
        with (
            files("price_updater.images.asset_logo")
            .joinpath(name + ".png")
            .open("rb") as logo_path
        ):
            return str(logo_path.name)

    def update(self, coins: List[Asset]) -> bool:
        try:
            workbook: Workbook = self.try_load_workbook()
            if workbook is not None:
                data = workbook["data"]
                i: int = 1
                while data.cell(row=1, column=i).value is not None:
                    if data.cell(row=1, column=i).value != "-":
                        price = next(
                            self.get_price(coin)
                            for coin in coins
                            if coin.name == data.cell(row=1, column=i).value
                        )
                        sheet = workbook[data.cell(row=2, column=i).value]
                        if price != "0,0":
                            sheet[data.cell(row=3, column=i).value] = price
                    i += 1
                workbook.save(settings.get_xlsx_file_path())
                return True
            return False
        except UnboundLocalError:
            return False

    def get_price(self, coin: Asset) -> str:
        match coin.chosen_currency:
            case Currency.USD:
                return coin.price_usd
            case Currency.PLN:
                return coin.price_pln
            case Currency.EUR:
                return coin.price_eur
            case Currency.GBP:
                return coin.price_gbp
        return "0,0"

    def get_asset_price(self, ticker: str) -> dict[Currency, str]:
        fiat_assets: Tuple[str, ...] = ("eur", "gbp", "usd")
        metal_assets: Tuple[str, ...] = ("xau", "xag")
        etf_assets: Tuple[str, ...] = ("ishares-swda", "ishares-emim")

        if ticker in fiat_assets:
            return self.get_fiat_price(ticker)
        if ticker in metal_assets:
            return self.get_metal_price(ticker)
        if ticker in etf_assets:
            return self.get_etf_price(ticker)
        return self.get_crypto_price(ticker)

    def try_load_workbook(self) -> Workbook | None:
        try:
            workbook: Workbook = load_workbook(settings.get_xlsx_file_path())
            return workbook
        except InvalidFileException:
            print("we need xlsx file!")
        except KeyError:
            print("please check xlsx format file!")
        except FileNotFoundError:
            print("file missing :D")
        return None

    def _exception_catch(self, error: Any) -> float:
        if isinstance(error, (exceptions.ConnectionError, exceptions.ReadTimeout)):
            currency.connection_lost = True
        return 0.0

    def get_fiat_price(self, ticker: str) -> dict[Currency, str]:
        try:
            match ticker:
                case "usd":
                    return {
                        Currency.USD: self._price_to_str(
                            currency.usd_pln, Currency.PLN, Currency.USD
                        ),
                        Currency.PLN: self._price_to_str(
                            currency.usd_pln, Currency.PLN, Currency.PLN
                        ),
                        Currency.EUR: self._price_to_str(
                            currency.usd_pln, Currency.PLN, Currency.EUR
                        ),
                        Currency.GBP: self._price_to_str(
                            currency.usd_pln, Currency.PLN, Currency.GBP
                        ),
                        Currency.LOGO: self.get_logo_path(ticker),
                    }
                case "gbp":
                    return {
                        Currency.USD: self._price_to_str(
                            currency.gbp_pln, Currency.PLN, Currency.USD
                        ),
                        Currency.PLN: self._price_to_str(
                            currency.gbp_pln, Currency.PLN, Currency.PLN
                        ),
                        Currency.EUR: self._price_to_str(
                            currency.gbp_pln, Currency.PLN, Currency.EUR
                        ),
                        Currency.GBP: self._price_to_str(
                            currency.gbp_pln, Currency.PLN, Currency.GBP
                        ),
                        Currency.LOGO: self.get_logo_path(ticker),
                    }
                case "eur":
                    return {
                        Currency.USD: self._price_to_str(
                            currency.eur_pln, Currency.PLN, Currency.USD
                        ),
                        Currency.PLN: self._price_to_str(
                            currency.eur_pln, Currency.PLN, Currency.PLN
                        ),
                        Currency.EUR: self._price_to_str(
                            currency.eur_pln, Currency.PLN, Currency.EUR
                        ),
                        Currency.GBP: self._price_to_str(
                            currency.eur_pln, Currency.PLN, Currency.GBP
                        ),
                        Currency.LOGO: self.get_logo_path(ticker),
                    }
                case _:
                    return {
                        Currency.USD: "0,0",
                        Currency.PLN: "0,0",
                        Currency.EUR: "0,0",
                        Currency.GBP: "0,0",
                        Currency.LOGO: self.get_logo_path(ticker),
                    }
        except ZeroDivisionError:
            return {
                Currency.USD: "0,0",
                Currency.PLN: "0,0",
                Currency.EUR: "0,0",
                Currency.GBP: "0,0",
                Currency.LOGO: self.get_logo_path(ticker),
            }

    def get_metal_price(self, ticker: str) -> dict[Currency, str]:
        def get_price() -> float:
            match ticker:
                case "xau":
                    local_ticker = "gold"
                case "xag":
                    local_ticker = "silver"
            try:
                url = f"https://markets.businessinsider.com/commodities/{local_ticker}-price"
                page = get(url, timeout=7)
                page_content = BeautifulSoup(page.content, "html.parser")
                onpage = str(
                    page_content.find("span", class_="price-section__current-value")
                )
                price_str = onpage.replace(
                    '<span class="price-section__current-value">', ""
                )
                price_str = price_str.replace("</span>", "")
                return float(price_str)
            except (
                ValueError,
                ZeroDivisionError,
                TypeError,
                exceptions.ConnectionError,
                exceptions.ReadTimeout,
            ) as error:
                value = self._exception_catch(error)
                return value

        price_exact: float = get_price()
        try:
            return {
                Currency.USD: self._price_to_str(
                    price_exact, Currency.USD, Currency.USD
                ),
                Currency.PLN: self._price_to_str(
                    price_exact, Currency.USD, Currency.PLN
                ),
                Currency.EUR: self._price_to_str(
                    price_exact, Currency.USD, Currency.EUR
                ),
                Currency.GBP: self._price_to_str(
                    price_exact, Currency.USD, Currency.GBP
                ),
                Currency.LOGO: self.get_logo_path(ticker),
            }
        except ZeroDivisionError:
            return {
                Currency.USD: "0,0",
                Currency.PLN: "0,0",
                Currency.EUR: "0,0",
                Currency.GBP: "0,0",
                Currency.LOGO: self.get_logo_path(ticker),
            }

    def get_etf_price(self, ticker: str) -> dict[Currency, str]:
        link: str
        if ticker == "ishares-swda":
            link = "https://markets.ft.com/data/etfs/tearsheet/summary?s=SWDA:LSE:GBX"
        else:
            link = "https://markets.ft.com/data/etfs/tearsheet/summary?s=EMIM:LSE:GBX"

        def get_price() -> float:
            try:
                page = get(link, timeout=7)
                page_content = BeautifulSoup(page.content, "html.parser")
                for onpage in page_content.find(
                    "span", class_="mod-ui-data-list__value"
                ):
                    page_str = str(onpage)
                    page_str = page_str.replace(",", "")
                    price = page_str[0:4]
                    return float(price)
            except (
                ValueError,
                ZeroDivisionError,
                TypeError,
                exceptions.ConnectionError,
                exceptions.ReadTimeout,
            ) as error:
                value = self._exception_catch(error)
                return value
            return 0.0

        price_exact = get_price()
        return {
            Currency.USD: self._price_to_str(price_exact, Currency.GBP, Currency.USD),
            Currency.PLN: self._price_to_str(price_exact, Currency.GBP, Currency.PLN),
            Currency.EUR: self._price_to_str(price_exact, Currency.GBP, Currency.EUR),
            Currency.GBP: self._price_to_str(price_exact, Currency.GBP, Currency.GBP),
            Currency.LOGO: self.get_logo_path(ticker),
        }

    def get_crypto_price(self, ticker: str) -> dict[Currency, str]:
        if ticker[-4:] == "_api":
            ticker = ticker.replace(ticker[-4:], "")
        url: str = f"https://coinmarketcap.com/currencies/{ticker.lower()}"
        page: Response = get(url, timeout=7)
        page_content = BeautifulSoup(page.content, "html.parser")

        price_float: float = 0.0

        def get_logo() -> str:
            try:
                logo_data = page_content.find("div", class_="sc-65e7f566-0 kYcmYb")
                logo_data = logo_data.find("img", src=True)
                return logo_data["src"]
            except (exceptions.ConnectionError, AttributeError):
                pass
            return "./images/asset_logo/none.png"

        def get_price_from_api() -> float:
            try:
                url_api: str = "https://api.coingecko.com/api/v3/simple/price"
                params: dict[str, str] = {"ids": ticker, "vs_currencies": "USD"}

                response: Response = get(url_api, params=params, timeout=5)
                if response.status_code == 200:
                    data = response.json()
                    price = data[ticker]["usd"]
                    return float(price)
                return 0.0
            except (
                AttributeError,
                ValueError,
                ZeroDivisionError,
                TypeError,
                KeyError,
                exceptions.ConnectionError,
                exceptions.ReadTimeout,
            ) as error:
                self._exception_catch(error)
                return 0.0

        def get_price() -> float:
            try:
                web = str(
                    page_content.find("span", class_="sc-65e7f566-0 WXGwg base-text")
                )
                price_str = web.replace(
                    '<span class="sc-65e7f566-0 WXGwg base-text" '
                    'data-test="text-cdp-price-display">$',
                    "",
                )
                price_str = price_str.replace("</span>", "")
                price_str = price_str.replace(",", "")
                price = float(price_str)
                return price
            except (
                AttributeError,
                ValueError,
                ZeroDivisionError,
                TypeError,
                KeyError,
                exceptions.ConnectionError,
                exceptions.ReadTimeout,
            ) as error:
                self._exception_catch(error)
                return 0.0

        price_float = 0.0
        if self.api_status == "api-off":
            price_float = get_price()
            if price_float == 0.0:
                price_float = get_price_from_api()
        else:
            i = 0
            while i < 4 and price_float == 0.0:
                sleep(0.2)
                price_float = get_price_from_api()
                i += 1
            if price_float == 0.0:
                price_float = get_price()

        if price_float == 0.0:
            return {
                Currency.USD: "0,0",
                Currency.PLN: "0,0",
                Currency.EUR: "0,0",
                Currency.GBP: "0,0",
                Currency.LOGO: get_logo(),
            }
        return {
            Currency.USD: self._price_to_str(
                price_float, Currency.USD, Currency.USD, True
            ),
            Currency.PLN: self._price_to_str(
                price_float, Currency.USD, Currency.PLN, True
            ),
            Currency.EUR: self._price_to_str(
                price_float, Currency.USD, Currency.EUR, True
            ),
            Currency.GBP: self._price_to_str(
                price_float, Currency.USD, Currency.GBP, True
            ),
            Currency.LOGO: get_logo(),
        }

    def _price_to_str(
        self,
        price: float,
        convert_from: Currency,
        convert_to: Currency,
        is_crypto: bool = False,
    ) -> str:
        price_float = 0.0
        if convert_from == Currency.PLN:
            match convert_to:
                case Currency.USD:
                    price_float = 1.0
                case Currency.EUR:
                    price_float = price / currency.eur_pln
                case Currency.GBP:
                    price_float = price / currency.gbp_pln
                case Currency.PLN:
                    price_float = price
        if convert_from == Currency.USD:
            match convert_to:
                case Currency.USD:
                    price_float = price
                case Currency.EUR:
                    price_float = price * (
                        currency.return_price(Currency.USD)
                        / currency.return_price(Currency.EUR)
                    )
                case Currency.GBP:
                    price_float = price * (
                        currency.return_price(Currency.USD)
                        / currency.return_price(Currency.GBP)
                    )
                case Currency.PLN:
                    price_float = price * currency.return_price(Currency.USD)
        elif convert_from == Currency.GBP:
            match convert_to:
                case Currency.USD:
                    price_float = price * (currency.gbp_pln / currency.usd_pln)
                case Currency.EUR:
                    price_float = price * (currency.gbp_pln / currency.eur_pln)
                case Currency.GBP:
                    price_float = price
                case Currency.PLN:
                    price_float = price * currency.gbp_pln
        if not is_crypto:
            price_str = f"{price_float:.3f}"
        else:
            price_str = f"{price_float:.10f}"
        price_str = price_str.rstrip("0")
        if price_str.endswith("."):
            price_str = price_str[:-1]
        price_str = price_str.replace(".", ",")
        return price_str
