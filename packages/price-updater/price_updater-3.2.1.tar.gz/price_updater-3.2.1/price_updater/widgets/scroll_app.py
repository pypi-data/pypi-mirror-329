import subprocess
from typing import List, Any
from kivymd.uix.scrollview import MDScrollView
from kivymd.uix.gridlayout import MDGridLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.clock import Clock
from openpyxl import load_workbook

from price_updater.widgets.coin_button import CoinButton
from price_updater.widgets.version_available import VersionAvailableMenu
from price_updater.lib.asset import Asset
from price_updater.lib.language import language, Text
from price_updater.lib.currency import currency, Currency
from price_updater.lib.settings import settings
from price_updater.lib.update import Update
from price_updater.lib.config import (
    font_config,
    color_orange_theme,
    color_button,
    color_error,
    color_asset_button,
    color_behind_window,
    color_window,
)


class ScrollApp(MDScrollView):
    spacing = 2
    coin_height = 40

    def __init__(self) -> None:
        super().__init__()
        self.coins_tab: List[Asset] = []
        self.coins = MDGridLayout(cols=1, spacing=self.spacing, size_hint_y=None)
        self.empty_list = Label(
            text=language.get_text(Text.EMPTY_LIST_TEXT.value),
            font_name=font_config,
            font_size=21,
            color=color_orange_theme,
        )
        self.loading_list = Label(
            text=language.get_text(Text.LOADING_LIST_TEXT.value),
            font_name=font_config,
            font_size=21,
            color=color_orange_theme,
        )
        self.add_widget(self.loading_list)
        self.coins.height = self.spacing + self.coin_height * len(self.coins_tab)
        self.bar_color = color_orange_theme
        self.bar_width = 5
        self.fetch_error: bool = False
        self.latest_version = self.newer_version()
        if self.latest_version:
            Clock.schedule_interval(self.install_version, 0.5)
        else:
            Clock.schedule_interval(self.show_coins, 2)

    def newer_version(self) -> str:
        installed_version: str = ""
        latest_version: str = ""

        try:
            versions_result = subprocess.run(
                ["pip", "index", "versions", "price_updater"],
                capture_output=True,
                text=True,
                check=False,
            )
        except:
            pass
        for line in versions_result.stdout.splitlines():
            if "INSTALLED:" in line:
                installed_version = line.split(":")[1].strip()
            elif "LATEST:" in line:
                latest_version = line.split(":")[1].strip()

        if installed_version and latest_version:
            if installed_version != latest_version:
                return latest_version
            return ""
        return ""

    def refresh_assets(self, instance: Any) -> None:
        self.coins.clear_widgets()
        self.clear_widgets()
        self.show_coins(self)

    def show_coins(self, instance: Any) -> None:
        Clock.unschedule(self.show_coins)
        currency.usd_pln = currency.get_currency(Currency.USD)
        currency.eur_pln = currency.get_currency(Currency.EUR)
        currency.gbp_pln = currency.get_currency(Currency.GBP)
        self.coins_tab = self.get_coins_from_xlsx()
        self.initialize_coins(check_fetch=True)
        self.clear_widgets()
        self.add_widget(self.coins)

    def initialize_coins(self, check_fetch: bool = False) -> None:
        self.coins.height = ScrollApp.spacing + ScrollApp.coin_height * len(
            self.coins_tab
        )
        self.coins.clear_widgets()
        i = 1
        if self.coins_tab:
            self.coins.add_widget(BoxLayout(size_hint=(1, 0.005)))
            for coin in self.coins_tab:
                coin_button = CoinButton(scrollapp=self, coin=coin)
                coin_button.text_color = color_button
                coin_button.md_bg_color = color_asset_button
                i += 1
                self.coins.add_widget(coin_button)
        else:
            self.coins.add_widget(self.empty_list)
            self.coins.height = self.spacing + self.coin_height * len(self.coins_tab)
        if check_fetch:
            self.fetch_error_msg()

    def get_coins_from_xlsx(self) -> list[Asset]:
        self.fetch_error = False
        try:
            workbook = load_workbook(settings.get_xlsx_file_path())
        except:
            return []

        if "data" not in workbook.sheetnames:
            workbook.create_sheet("data")
            hidden = workbook["data"]
            hidden.sheet_state = "hidden"
            workbook.save(settings.get_xlsx_file_path())
        data = workbook["data"]

        coins: List[Asset] = []

        i = 1
        while data.cell(row=1, column=i).value is not None:
            if data.cell(row=1, column=i).value != "-":
                ticker = data.cell(row=1, column=i).value
                worksheet = data.cell(row=2, column=i).value
                cell = data.cell(row=3, column=i).value
                curr_currency = data.cell(row=4, column=i).value
                price = Update().get_asset_price(ticker)
                try:
                    if price[Currency.PLN] != "0,0":
                        coins.append(
                            Asset(
                                asset_id=i,
                                name=ticker,
                                worksheet=worksheet,
                                cell=cell,
                                price=price,
                                currency=Currency(curr_currency),
                            )
                        )
                    else:
                        self.fetch_error = True
                        coins.append(
                            Asset(
                                asset_id=i,
                                name=ticker,
                                worksheet=worksheet,
                                cell=cell,
                                price={
                                    Currency.USD: "0,0",
                                    Currency.PLN: "0,0",
                                    Currency.EUR: "0,0",
                                    Currency.GBP: "0,0",
                                    Currency.LOGO: price[Currency.LOGO],
                                },
                                currency=Currency(curr_currency),
                            )
                        )
                except ValueError:
                    self.fetch_error = True
                    coins.append(
                        Asset(
                            asset_id=i,
                            name=ticker,
                            worksheet=worksheet,
                            cell=cell,
                            price={
                                Currency.USD: "0,0",
                                Currency.PLN: "0,0",
                                Currency.EUR: "0,0",
                                Currency.GBP: "0,0",
                                Currency.LOGO: price[Currency.LOGO],
                            },
                            currency=Currency(curr_currency),
                        )
                    )
            i += 1
        return coins

    def try_fetch_data_connection_lost(self, instance: Any) -> None:
        test_value: float = currency.get_currency(Currency.USD)
        if test_value != 0.0:
            Clock.unschedule(self.try_fetch_data_connection_lost)
            currency.connection_lost = False
            self.show_coins(self)

    def install_version(self, instance: Any) -> None:
        outdated_version = Popup(
            title_color=color_orange_theme,
            overlay_color=color_behind_window,
            separator_color=color_orange_theme,
            size_hint=(None, None),
            size=(350, 200),
            auto_dismiss=False,
            title_font=font_config,
            title=language.get_text(Text.OUTDATED_VERSION.value),
            background_color=color_window,
        )

        version_menu = VersionAvailableMenu(self, outdated_version, self.latest_version)
        outdated_version.content = version_menu
        outdated_version.open(animation=True)

    def fetch_error_msg(self) -> None:
        if currency.connection_lost:
            Clock.schedule_interval(self.try_fetch_data_connection_lost, 5)
        if self.fetch_error:
            warning_msg = Popup(
                title_color=color_orange_theme,
                overlay_color=color_behind_window,
                separator_color=color_orange_theme,
                size_hint=(None, None),
                size=(350, 200),
                auto_dismiss=True,
                title_font=font_config,
                title=language.get_text(Text.FETCH_ERROR_TITLE.value),
                background_color=color_window,
            )
            warning_content = Label(
                text=(
                    language.get_text(Text.CONNECTION_LOST.value)
                    if currency.connection_lost
                    else language.get_text(Text.FETCH_ERROR_MSG.value)
                ),
                font_name=font_config,
                color=color_error,
            )
            warning_msg.content = warning_content
            warning_msg.open(animation=True)
