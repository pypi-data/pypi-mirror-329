import re
from typing import Any, Dict
from dataclasses import dataclass
from openpyxl.workbook import Workbook
from kivymd.uix.button import MDRaisedButton
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.checkbox import CheckBox
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label

from price_updater.widgets.scroll_app import ScrollApp
from price_updater.lib.on_check_box_active import OnCheckBoxActive
from price_updater.lib.update import Update
from price_updater.lib.asset import Asset
from price_updater.lib.currency import Currency
from price_updater.lib.text_input import TextInputC
from price_updater.lib.button import ButtonC
from price_updater.lib.language import language, Text
from price_updater.lib.settings import settings
from price_updater.lib.auto_suggestion_text import AutoSuggestionText
from price_updater.lib.config import (
    font_config,
    color_orange_theme,
    color_button,
    color_checkbox,
    color_error,
)
from price_updater.coins_list import assets_list


@dataclass
class Info:
    check: bool
    price: Dict[Currency, str]


class AddMenu(BoxLayout, OnCheckBoxActive):
    def __init__(self, scrollApp: ScrollApp, popup: Popup, **kwargs: Any) -> None:
        super().__init__(**kwargs)
        self.scrollapp: ScrollApp = scrollApp
        self.popup: Popup = popup
        self.orientation: str = "vertical"
        self.opacity: float = 0.8
        self.spacing: int = 5
        self.workbook: Workbook = Update().try_load_workbook()
        self.worksheet_input: str = ""

        if self.workbook is not None:
            self.build()
        else:
            self.no_workbook_label = Label(
                text=language.get_text(Text.PLEASE_SELECT_WORKBOOK.value),
                font_name=font_config,
            )
            self.button_ok = ButtonC(
                text="OK",
                on_release=self.popup.dismiss,
                size_hint=(1, 0.2),
                text_color=color_orange_theme,
            )
            self.add_widget(self.no_workbook_label)
            self.add_widget(self.button_ok)

    def build(self) -> None:
        self.sheets: list[str] = self.workbook.sheetnames
        self.sheets.remove("data")

        self.scroll_sheets = ScrollView()
        self.sheets_widget = BoxLayout(
            orientation="vertical", size_hint_y=None, spacing=2
        )
        self.sheets_widget.bind(minimum_height=self.sheets_widget.setter("height"))
        self.scroll_sheets.add_widget(self.sheets_widget)
        i = 0
        for sheet in self.sheets:
            sheet_button = MDRaisedButton(
                text=sheet,
                md_bg_color=color_button,
                size_hint=(1, None),
                height=35,
                on_release=self.chosen_sheet,
                font_name=font_config,
                font_size=17,
                text_color=color_orange_theme,
            )
            if i == 0:
                self.worksheet_input = sheet
                sheet_button.md_bg_color = color_orange_theme
                sheet_button.text_color = color_button
            self.sheets_widget.add_widget(sheet_button)
            i += 1
        self.coin_name_input = AutoSuggestionText(
            text=language.get_text(Text.COIN_NAME.value), suggestions=assets_list
        )
        self.coin_name_input.select_all()
        self.coin_name_input.focus = True
        self.cell_input = TextInputC(text=language.get_text(Text.CELL.value))
        self.checkboxes_currency = BoxLayout(
            orientation="horizontal", size_hint=(1, 0.15)
        )
        self.checkbox_currency_labels = BoxLayout(
            orientation="horizontal", size_hint=(1, 0.07)
        )
        self.checkbox_usd = CheckBox(active=True, color=color_checkbox)
        self.checkbox_usd.bind(active=self.on_checkbox_active)
        self.checkbox_eur = CheckBox(active=False, color=color_checkbox)
        self.checkbox_eur.bind(active=self.on_checkbox_active)
        self.checkbox_gbp = CheckBox(active=False, color=color_checkbox)
        self.checkbox_gbp.bind(active=self.on_checkbox_active)
        self.checkbox_pln = CheckBox(active=False, color=color_checkbox)
        self.checkbox_pln.bind(active=self.on_checkbox_active)
        self.label_usd = Label(
            text="USD", color=color_orange_theme, font_name=font_config, font_size=17
        )
        self.label_eur = Label(
            text="EUR", color=color_orange_theme, font_name=font_config, font_size=17
        )
        self.label_gbp = Label(
            text="GBP", color=color_orange_theme, font_name=font_config, font_size=17
        )
        self.label_pln = Label(
            text="PLN", color=color_orange_theme, font_name=font_config, font_size=17
        )
        self.checkbox_currency_labels.add_widget(self.label_usd)
        self.checkbox_currency_labels.add_widget(self.label_eur)
        self.checkbox_currency_labels.add_widget(self.label_gbp)
        self.checkbox_currency_labels.add_widget(self.label_pln)
        self.checkboxes_currency.add_widget(self.checkbox_usd)
        self.checkboxes_currency.add_widget(self.checkbox_eur)
        self.checkboxes_currency.add_widget(self.checkbox_gbp)
        self.checkboxes_currency.add_widget(self.checkbox_pln)

        self.add_widget(self.coin_name_input)
        self.add_widget(self.scroll_sheets)
        self.add_widget(self.cell_input)

        self.add_widget(self.checkboxes_currency)
        self.add_widget(self.checkbox_currency_labels)

        buttons = BoxLayout(orientation="horizontal", size_hint=(1, 0.4))
        self.add_widget(buttons)
        button_add = ButtonC(
            text=language.get_text(Text.ADD.value),
            on_release=self.add_this_coin,
            size_hint=(1, 0.8),
        )
        buttons.add_widget(button_add)

    def chosen_sheet(self, instance: MDRaisedButton) -> None:
        if instance.md_bg_color == color_orange_theme:
            self.worksheet_input = instance.text
            for sheet in self.sheets_widget.children:
                if instance is not sheet:
                    sheet.md_bg_color = color_button
                    sheet.text_color = color_orange_theme
        else:
            for sheet in self.sheets_widget.children:
                sheet.md_bg_color = color_button
                sheet.text_color = color_orange_theme
            self.worksheet_input = instance.text
            instance.md_bg_color = color_orange_theme
            instance.text_color = color_button

    def add_this_coin(self, instance: ButtonC) -> None:
        info = self.check_input_data()
        if info.check:
            data = self.workbook["data"]
            i = 1
            while data.cell(row=1, column=i).value not in ("-", None):
                i += 1

            chosen_currency: Currency
            if self.checkbox_usd.active:
                chosen_currency = Currency.USD
            elif self.checkbox_eur.active:
                chosen_currency = Currency.EUR
            elif self.checkbox_gbp.active:
                chosen_currency = Currency.GBP
            else:
                chosen_currency = Currency.PLN

            data.cell(row=1, column=i).value = self.coin_name_input.text.lower()
            data.cell(row=2, column=i).value = self.worksheet_input
            data.cell(row=3, column=i).value = self.cell_input.text.upper()
            data.cell(row=4, column=i).value = chosen_currency.name
            self.workbook.save(settings.get_xlsx_file_path())
            self.scrollapp.coins_tab.append(
                Asset(
                    asset_id=i,
                    name=self.coin_name_input.text.lower(),
                    worksheet=self.worksheet_input,
                    cell=self.cell_input.text.upper(),
                    price=info.price,
                    currency=chosen_currency,
                )
            )
            self.scrollapp.initialize_coins()
            self.scrollapp.coins.height = (
                ScrollApp.spacing
                + ScrollApp.coin_height * len(self.scrollapp.coins_tab)
            )
            self.popup.dismiss()

    def check_input_data(self) -> Info:
        test_price: dict[Currency, str]
        if (
            self.coin_name_input.text != language.get_text(Text.COIN_NAME.value)
            and self.coin_name_input.text != ""
        ):
            test_price = Update().get_asset_price(self.coin_name_input.text)
        else:
            test_price = {
                Currency.USD: "0,0",
                Currency.PLN: "0,0",
                Currency.GBP: "0,0",
                Currency.EUR: "0,0",
                Currency.LOGO: "",
            }

        name_ok: bool = False
        sheet_ok: bool = False
        cell_ok: bool = False
        currency_ok: bool = False
        ##############################################
        if test_price is not None and test_price[Currency.EUR] != "0,0":
            self.coin_name_input.text_ok()
            name_ok = True
        else:
            self.coin_name_input.text_error()
        ##############################################
        if "data" not in self.workbook.sheetnames:
            self.workbook.create_sheet("data")
            hidden = self.workbook["data"]
            hidden.sheet_state = "hidden"
            self.workbook.save(settings.get_xlsx_file_path())
        if self.worksheet_input != "":
            sheet_ok = True
        else:
            for sheet in self.sheets_widget.children:
                sheet.color = color_error
        ##############################################
        cell_pattern = r"^[A-Za-z]\d+$"
        if re.match(cell_pattern, self.cell_input.text):
            self.cell_input.text_ok()
            cell_ok = True
        else:
            self.cell_input.text_error()
        ##############################################
        if (
            self.checkbox_usd.active
            or self.checkbox_eur.active
            or self.checkbox_gbp.active
            or self.checkbox_pln.active
        ):
            currency_ok = True

        ##############################################
        return Info(check=name_ok & sheet_ok & cell_ok & currency_ok, price=test_price)
