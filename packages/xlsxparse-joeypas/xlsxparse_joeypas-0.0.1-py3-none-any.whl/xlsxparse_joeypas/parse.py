from openpyxl import load_workbook
import re

def find_cell_references(file_path, sheet_name=None):
    wb = load_workbook(file_path, data_only=False)
    sheet = wb[sheet_name] if sheet_name else wb.active

    cell_refrences = {}

    formula_pattern = re.compile(r'([A-Z]+\d+)')

    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                referenced_cells = formula_pattern.findall(cell.value)
                cell_refrences[cell.coordinate] = referenced_cells

    return cell_refrences

